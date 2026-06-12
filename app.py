# ================================================
# EDUTRACK AI — ARQUIVO PRINCIPAL
# ================================================
# ESTRUTURA DO ARQUIVO:
#   1. Importações
#   2. Configuração da API Xano
#   3. *** CSS DO DESIGN (st.markdown) ***  <-- AQUI FICA O VISUAL
#   4. Funções de conexão e utilitários
#   5. Sistema de autenticação
#   6. Módulos CRUD (Professores, Disciplinas, Tarefas)
#   7. Gerador de PDF
#   8. Dashboard
#   9. Navegação principal
# ================================================

# ------------------------------------------------
# 1. IMPORTAÇÃO DE BIBLIOTECAS
# ------------------------------------------------

import io                          # Manipulação de arquivos em memória (PDF/Excel)
from datetime import date          # Data atual para relatórios e cálculos de prazo

import pandas as pd                # Manipulação de dados em tabelas (DataFrames)
import plotly.express as px        # Gráficos interativos (barras, linhas, pizza)
import requests                    # Chamadas HTTP para a API do Xano
import streamlit as st             # Framework principal da interface web
from reportlab.lib import colors   # Cores para o PDF
from reportlab.lib.pagesizes import A4  # Tamanho de página A4 para o PDF
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet  # Estilos de texto no PDF
from reportlab.lib.units import cm # Unidade centímetro para medidas no PDF
from reportlab.platypus import (   # Componentes de layout do PDF:
    Paragraph,                     #   - Parágrafos de texto
    SimpleDocTemplate,             #   - Template do documento
    Spacer,                        #   - Espaços em branco
    Table,                         #   - Tabelas
    TableStyle,                    #   - Estilos de tabela
)

# ------------------------------------------------
# 2. CONFIGURAÇÃO DA API XANO
# ------------------------------------------------

# Tenta carregar a URL da API do arquivo secrets.toml (para deploy)
# Se não encontrar (ambiente local), usa a URL hardcoded
try:
    BASE_URL = st.secrets["xano"]["base_url"]
except Exception:
    BASE_URL = "https://x8ki-letl-twmt.n7.xano.io/api:aya_6V8d"

# ================================================
# 3. CSS DO DESIGN — COLOCADO AQUI NO TOPO
# ================================================
# IMPORTANTE: este bloco DEVE ficar antes de qualquer
# chamada a st.title(), st.header(), st.sidebar etc.
# O st.set_page_config() já foi chamado mais abaixo,
# mas o CSS é injetado aqui mesmo assim pois o
# Streamlit aplica o HTML ao carregar a página.
#
# COMO FUNCIONA:
#   - st.markdown(..., unsafe_allow_html=True) injeta
#     HTML/CSS diretamente no <head> da página.
#   - Usamos seletores CSS do Streamlit para sobrescrever
#     os estilos padrão sem precisar de bibliotecas externas.
#   - Toda a identidade visual (cores, fontes, espaçamentos)
#     está concentrada aqui — fácil de editar depois.
# ================================================

def aplicar_css():
    """Injeta CSS de dark/light mode conforme session_state.tema."""
    # Tema fixo: dark
    v = {
            'bg':         '#0f1117',
            'surface':    '#1a1d27',
            'input':      '#0d0f18',
            'border':     'rgba(255,255,255,0.08)',
            'border_h':   'rgba(255,255,255,0.18)',
            'accent':     '#f5c842',
            'accent_dim': 'rgba(245,200,66,0.12)',
            'accent_txt': '#6b4f00',
            'accent_drk': '#f5c842',
            'green':      '#34d399',
            'green_dim':  'rgba(52,211,153,0.12)',
            'amber':      '#fbbf24',
            'amber_dim':  'rgba(251,191,36,0.12)',
            'red':        '#f87171',
            'red_dim':    'rgba(248,113,113,0.12)',
            'purple':     '#a78bfa',
            'purple_dim': 'rgba(167,139,250,0.12)',
            'text':       '#f1f5f9',
            'muted':      '#94a3b8',
            'dim':        '#64748b',
            'shadow':     'none',
    }

    css = """
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap');
:root {{
  --bg: {bg}; --surface: {surface}; --input: {input};
  --border: {border}; --border-h: {border_h};
  --accent: {accent}; --accent-dim: {accent_dim};
  --accent-txt: {accent_txt}; --accent-drk: {accent_drk};
  --green: {green}; --green-dim: {green_dim};
  --amber: {amber}; --amber-dim: {amber_dim};
  --red: {red}; --red-dim: {red_dim};
  --purple: {purple}; --purple-dim: {purple_dim};
  --text: {text}; --muted: {muted}; --dim: {dim};
  --shadow: {shadow}; --radius: 10px;
}}
html,body,[class*="css"],.stApp{{font-family:'Inter',sans-serif!important}}
.stApp{{background:var(--bg)!important}}
p,span,label,div{{color:var(--text)}}
[data-testid="stSidebar"]{{background:var(--surface)!important;border-right:0.5px solid var(--border)!important;box-shadow:var(--shadow)!important}}
[data-testid="stSidebar"]>div:first-child{{background:var(--surface)!important}}
[data-testid="stSidebar"] *{{color:var(--muted)!important}}
[data-testid="stSidebar"] h1{{font-size:16px!important;font-weight:600!important;color:var(--text)!important;padding-bottom:12px!important;border-bottom:0.5px solid var(--border)!important;margin-bottom:6px!important}}
[data-testid="stSidebar"] .stRadio label:first-child{{font-size:10px!important;text-transform:uppercase!important;letter-spacing:.1em!important;color:var(--dim)!important}}
[data-testid="stSidebar"] .stRadio div[role="radiogroup"] label{{font-size:13px!important;color:var(--muted)!important;padding:8px 12px!important;border-radius:8px!important;border-left:3px solid transparent!important;white-space:nowrap!important}}
[data-testid="stSidebar"] .stRadio div[role="radiogroup"] label:has(input:checked){{background:var(--accent-dim)!important;color:var(--accent-drk)!important;border-left-color:var(--accent)!important;font-weight:600!important}}
[data-testid="stSidebar"] .stRadio div[role="radiogroup"] p{{color:inherit!important}}
[data-testid="stSidebar"] .stButton button{{background:transparent!important;border:0.5px solid var(--border-h)!important;color:var(--muted)!important;font-size:13px!important;border-radius:8px!important;padding:8px 16px!important;width:100%!important}}
[data-testid="stSidebar"] .stButton button:hover{{background:var(--red-dim)!important;border-color:var(--red)!important;color:var(--red)!important}}
h1{{font-size:22px!important;font-weight:600!important;color:var(--text)!important;padding-bottom:10px!important;border-bottom:0.5px solid var(--border)!important;margin-bottom:18px!important}}
h2{{font-size:16px!important;font-weight:600!important;color:var(--text)!important;margin-top:18px!important}}
h3{{font-size:14px!important;font-weight:500!important;color:var(--muted)!important}}
[data-testid="stMetric"]{{background:var(--surface)!important;border:0.5px solid var(--border)!important;border-radius:var(--radius)!important;padding:14px 18px!important;box-shadow:var(--shadow)!important}}
[data-testid="stMetricLabel"] p{{font-size:11px!important;color:var(--dim)!important;text-transform:uppercase!important;letter-spacing:.07em!important}}
[data-testid="stMetricValue"]{{font-size:24px!important;font-weight:600!important;color:var(--text)!important}}
.stButton button[kind="primary"]{{
  background:#1e2130!important;color:#f1f5f9!important;
  border:0.5px solid rgba(255,255,255,0.15)!important;
  border-radius:8px!important;
  font-size:13px!important;font-weight:500!important;padding:9px 20px!important
}}
.stButton button[kind="primary"]:hover{{
  background:#2a2f45!important;
  border-color:rgba(255,255,255,0.25)!important;
}}
.stButton button[kind="secondary"],.stButton button:not([kind]){{
  background:transparent!important;color:var(--muted)!important;
  border:0.5px solid var(--border-h)!important;
  border-radius:8px!important;font-size:13px!important;padding:9px 20px!important
}}
.stButton button:not([kind]):hover{{
  background:rgba(255,255,255,0.05)!important;
  border-color:rgba(255,255,255,0.2)!important;
}}
.stTextInput input,.stNumberInput input,.stTextArea textarea{{
  background:var(--input)!important;
  border:0.5px solid var(--border-h)!important;
  border-radius:8px!important;color:var(--text)!important;font-size:13px!important
}}
.stTextInput input:hover,.stNumberInput input:hover{{
  border-color:var(--accent)!important
}}
.stTextInput input:focus,.stNumberInput input:focus{{border-color:var(--accent)!important;box-shadow:0 0 0 2px var(--accent-dim)!important}}
.stTextInput label,.stNumberInput label,.stSelectbox label,.stMultiSelect label,.stCheckbox label,.stTextArea label{{font-size:11px!important;color:var(--dim)!important;font-weight:500!important;text-transform:uppercase!important;letter-spacing:.05em!important}}
.stSelectbox>div>div,.stMultiSelect>div>div{{background:var(--input)!important;border:0.5px solid var(--border-h)!important;border-radius:8px!important;color:var(--text)!important}}
.stMultiSelect span[data-baseweb="tag"]{{background:var(--accent-dim)!important;border:0.5px solid var(--accent)!important;border-radius:20px!important;color:var(--accent-drk)!important;font-size:12px!important}}
[data-testid="stExpander"]{{background:var(--surface)!important;border:0.5px solid var(--border)!important;border-radius:var(--radius)!important;margin-bottom:12px!important;box-shadow:var(--shadow)!important}}
[data-testid="stExpander"] summary{{font-size:13px!important;color:var(--muted)!important}}
[data-testid="stExpander"] summary:hover{{color:var(--accent-drk)!important}}
[data-testid="stDataFrame"],[data-testid="stDataEditor"]{{border:0.5px solid var(--border)!important;border-radius:var(--radius)!important;overflow:hidden!important}}
[data-testid="stDataFrame"] th,[data-testid="stDataEditor"] th{{background:var(--input)!important;color:var(--dim)!important;font-size:11px!important;text-transform:uppercase!important;letter-spacing:.07em!important;border-bottom:0.5px solid var(--border)!important}}
[data-testid="stDataFrame"] td,[data-testid="stDataEditor"] td{{background:var(--surface)!important;color:var(--muted)!important;font-size:13px!important;border-bottom:0.5px solid var(--border)!important}}
.stCheckbox input[type="checkbox"]{{accent-color:var(--accent)!important}}
.stTabs [data-baseweb="tab-list"]{{background:transparent!important;border-bottom:0.5px solid var(--border)!important;gap:4px!important}}
.stTabs [data-baseweb="tab"]{{font-size:13px!important;color:var(--muted)!important;background:transparent!important;border-bottom:2px solid transparent!important;padding:8px 16px!important}}
.stTabs [aria-selected="true"]{{color:var(--accent-drk)!important;border-bottom-color:var(--accent)!important;font-weight:600!important}}
[data-testid="stAlert"]{{border-radius:8px!important;font-size:13px!important}}
[data-testid="stAlert"][kind="success"]{{background:var(--green-dim)!important;border-left:3px solid var(--green)!important}}
[data-testid="stAlert"][kind="success"] p{{color:var(--green)!important}}
[data-testid="stAlert"][kind="error"]{{background:var(--red-dim)!important;border-left:3px solid var(--red)!important}}
[data-testid="stAlert"][kind="error"] p{{color:var(--red)!important}}
[data-testid="stAlert"][kind="warning"]{{background:var(--amber-dim)!important;border-left:3px solid var(--amber)!important}}
[data-testid="stAlert"][kind="warning"] p{{color:var(--amber)!important}}
[data-testid="stAlert"][kind="info"]{{background:var(--accent-dim)!important;border-left:3px solid var(--accent)!important}}
[data-testid="stAlert"][kind="info"] p{{color:var(--accent-drk)!important}}
[data-testid="stForm"]{{background:var(--surface)!important;border:0.5px solid var(--border)!important;border-radius:var(--radius)!important;padding:18px!important;box-shadow:var(--shadow)!important}}
[data-testid="stDownloadButton"] button{{background:var(--accent-dim)!important;border:0.5px solid var(--accent)!important;color:var(--accent-drk)!important;border-radius:8px!important;font-size:13px!important;font-weight:600!important}}
[data-testid="stCaptionContainer"] p{{color:var(--dim)!important;font-size:12px!important}}
hr{{border:none!important;border-top:0.5px solid var(--border)!important;margin:18px 0!important}}
::-webkit-scrollbar{{width:4px!important;height:4px!important}}
::-webkit-scrollbar-track{{background:transparent!important}}
::-webkit-scrollbar-thumb{{background:var(--accent)!important;border-radius:4px!important}}
#MainMenu{{visibility:hidden!important}}footer{{visibility:hidden!important}}header{{visibility:hidden!important}}
[data-testid="stSidebarCollapsedControl"]{{display:none!important}}
button[data-testid="baseButton-headerNoPadding"]{{display:none!important}}
section[data-testid="stSidebar"]{{
  display:block!important;
  min-width:244px!important;
  visibility:visible!important;
  transform:translateX(0)!important;
}}
section[data-testid="stSidebar"][aria-expanded="false"]{{
  display:block!important;
  transform:translateX(0)!important;
  min-width:244px!important;
}}

[data-testid="InputInstructions"]{{display:none!important}}
small.st-emotion-cache-ztfqz8{{display:none!important}}
.st-emotion-cache-ztfqz8{{display:none!important}}
[data-testid="stForm"] small{{display:none!important}}
.stTextInput small{{display:none!important}}
.stTextInput [data-testid="InputInstructions"]{{display:none!important}}
""".format(**v)

    st.markdown('<style>' + css + '</style>', unsafe_allow_html=True)

    # Bloqueia Enter para não submeter formulários.
    # st.markdown não executa scripts — usamos components.v1.html
    # que roda dentro de um iframe e acessa o documento pai via window.parent
    import streamlit.components.v1 as _components
    _components.html("""
    <script>
    (function() {
        var parent = window.parent ? window.parent.document : document;
        parent.addEventListener('keydown', function(e) {
            if (e.key === 'Enter' &&
                e.target && e.target.tagName === 'INPUT' &&
                e.target.type !== 'submit' &&
                e.target.type !== 'button') {
                e.stopPropagation();
                e.preventDefault();
            }
        }, true);
    })();
    </script>
    """, height=0)


# ------------------------------------------------
# 4. FUNÇÕES DE CONEXÃO E UTILITÁRIOS
# ------------------------------------------------

def get_headers():
    '''
    Gera o cabeçalho HTTP com o token JWT do usuário logado.
    O token é armazenado no session_state após o login e enviado
    em todas as requisições para o Xano identificar o usuário.
    Sem o token, o Xano rejeita a requisição (status 401).
    '''
    headers = {'Content-Type': 'application/json'}
    if 'auth_token' in st.session_state:
        # Bearer token — padrão JWT de autenticação
        headers['Authorization'] = f'Bearer {st.session_state.auth_token}'
    return headers


@st.cache_data(ttl=30, show_spinner=False)
def api_get(endpoint: str, _token: str = ''):
    '''
    Busca dados do Xano com cache inteligente de 30 segundos.

    O parâmetro _token (prefixo _ = ignorado pelo cache do Streamlit
    normalmente, mas aqui forçamos sua inclusão na chave de cache)
    garante que cada usuário tenha seu próprio cache isolado.
    Sem isso, usuário B veria dados do usuário A em cache.

    TTL de 30s = dados ficam em cache por 30 segundos antes de
    buscar novamente no Xano — reduz chamadas à API.
    '''
    try:
        headers = {'Content-Type': 'application/json'}
        if _token:
            headers['Authorization'] = f'Bearer {_token}'
        resposta = requests.get(f'{BASE_URL}/{endpoint}', headers=headers, timeout=10)
        if resposta.status_code == 200:
            return resposta.json()
        st.warning(f"Erro ao carregar '{endpoint}': status {resposta.status_code}")
        return []
    except requests.exceptions.RequestException as e:
        st.error(f'Falha de conexão com a API: {e}')
        return []


def _invalidar_cache():
    '''
    Limpa o cache do api_get para forçar recarga imediata dos dados.
    Chamado sempre após criar, editar ou deletar um registro,
    garantindo que a próxima leitura busque dados atualizados do Xano.
    '''
    api_get.clear()


def api_post(endpoint: str, dados: dict):
    '''
    Envia um POST para criar um novo registro no Xano.
    O Xano usa o token JWT no cabeçalho para associar o
    registro ao usuário logado (campo user_id automático).
    '''
    try:
        return requests.post(
            f'{BASE_URL}/{endpoint}', json=dados, headers=get_headers(), timeout=10
        )
    except requests.exceptions.RequestException as e:
        st.error(f'Erro ao salvar dados: {e}')
        return None


def api_patch(endpoint: str, id: int, dados: dict):
    '''
    Envia um PATCH para atualizar campos específicos de um registro.
    Diferente do PUT (que substitui tudo), o PATCH atualiza
    apenas os campos enviados — os demais permanecem inalterados.
    Ex: api_patch('tarefas', 5, {'concluida': True}) atualiza
    só o campo concluida da tarefa de id=5.
    '''
    try:
        return requests.patch(
            f'{BASE_URL}/{endpoint}/{id}', json=dados, headers=get_headers(), timeout=10
        )
    except requests.exceptions.RequestException as e:
        st.error(f'Erro ao atualizar dados: {e}')
        return None


def api_delete(endpoint: str, id: int):
    '''
    Envia um DELETE para remover permanentemente um registro do Xano.
    Usado com cascata: ao deletar professor, disciplinas, tarefas,
    metas e anotações relacionadas são deletadas primeiro.
    '''
    try:
        return requests.delete(
            f'{BASE_URL}/{endpoint}/{id}', headers=get_headers(), timeout=10
        )
    except requests.exceptions.RequestException as e:
        st.error(f'Erro ao remover registro: {e}')
        return None


# ------------------------------------------------
# 5. SISTEMA DE AUTENTICAÇÃO
# ------------------------------------------------

def tela_acesso():
    st.markdown("""
    <style>
    /* Fundo geral */
    .stApp { background: #0f1117 !important; }
    [data-testid="InputInstructions"] { display: none !important; }

    /* Coluna esquerda — fundo ligeiramente diferente */
    [data-testid="stColumns"] > div:first-child {
        background: #1a1d27;
        border-right: 0.5px solid rgba(255,255,255,0.07);
        padding: 40px 32px !important;
        min-height: 100vh;
    }

    /* Textos do painel esquerdo */
    [data-testid="stColumns"] > div:first-child h2 {
        font-size: 20px !important;
        font-weight: 500 !important;
        color: #f1f5f9 !important;
        margin-bottom: 8px !important;
    }
    [data-testid="stColumns"] > div:first-child p {
        color: #94a3b8 !important;
        font-size: 13px !important;
        line-height: 1.65 !important;
    }
    [data-testid="stColumns"] > div:first-child [data-testid="stCaptionContainer"] p {
        color: #334155 !important;
        font-size: 11px !important;
    }

    /* Coluna direita — formulário com padding */
    [data-testid="stColumns"] > div:last-child {
        padding: 48px 40px !important;
        background: #0f1117 !important;
    }

    /* Título e caption do formulário */
    [data-testid="stColumns"] > div:last-child h3 {
        font-size: 20px !important;
        font-weight: 500 !important;
        color: #f1f5f9 !important;
        margin-bottom: 4px !important;
    }
    [data-testid="stColumns"] > div:last-child [data-testid="stCaptionContainer"] p {
        color: #64748b !important;
        font-size: 13px !important;
        margin-bottom: 20px !important;
    }

    /* Tabs — aba ativa em amarelo */
    .stTabs [data-baseweb="tab-list"] {
        background: transparent !important;
        border-bottom: 0.5px solid rgba(255,255,255,0.08) !important;
        margin-bottom: 24px !important;
        gap: 4px !important;
    }
    .stTabs [data-baseweb="tab"] {
        font-size: 14px !important;
        color: #64748b !important;
        background: transparent !important;
        border-bottom: 2px solid transparent !important;
        padding: 8px 18px !important;
    }
    .stTabs [aria-selected="true"] {
        color: #f5c842 !important;
        border-bottom-color: #f5c842 !important;
        font-weight: 500 !important;
        background: transparent !important;
    }

    /* Inputs dark */
    .stTextInput input {
        background: #0d0f18 !important;
        border: 0.5px solid rgba(255,255,255,0.12) !important;
        border-radius: 8px !important;
        color: #f1f5f9 !important;
        font-size: 14px !important;
        padding: 11px 14px !important;
    }
    .stTextInput input:focus {
        border-color: #f5c842 !important;
        box-shadow: 0 0 0 2px rgba(245,200,66,0.15) !important;
    }
    .stTextInput label {
        font-size: 11px !important;
        font-weight: 500 !important;
        color: #64748b !important;
        text-transform: uppercase !important;
        letter-spacing: .06em !important;
    }

    /* Form sem borda */
    [data-testid="stForm"] {
        background: transparent !important;
        border: none !important;
        padding: 0 !important;
        box-shadow: none !important;
    }



    /* File uploader */
    [data-testid="stFileUploader"] {
        background: #0d0f18 !important;
        border: 1px dashed rgba(255,255,255,0.12) !important;
        border-radius: 8px !important;
    }
    [data-testid="stFileUploader"] label {
        color: #64748b !important;
        font-size: 13px !important;
    }
    </style>
    """, unsafe_allow_html=True)

    col_esq, col_dir = st.columns([1, 1.8])

    # ── Painel esquerdo ────────────────────────────────────
    with col_esq:
        st.markdown("## 📚 &nbsp; EduTrack.")
        st.markdown("### Seu portal acadêmico")
        st.markdown('<p style="color:#94a3b8;font-size:13px;line-height:1.65">Gerencie disciplinas, acompanhe suas notas e evolua com inteligência.</p>', unsafe_allow_html=True)
        st.markdown("---")
        st.markdown("📊 &nbsp; Dashboard de desempenho")
        st.markdown("📅 &nbsp; Calendário de provas")
        st.markdown("🎯 &nbsp; Metas por disciplina")
        st.markdown("📄 &nbsp; Relatório em PDF")
        st.markdown("---")
        st.caption("EduTrack AI · 2026")

    # ── Formulário ─────────────────────────────────────────
    with col_dir:
        # Padding interno para centralizar e estreitar o formulário
        _, col_f, _ = st.columns([1, 10, 1])

        with col_f:
            tab_login, tab_cadastro = st.tabs(["  Entrar  ", "  Criar conta  "])

            with tab_login:
                st.markdown("### Bem-vindo de volta")
                st.markdown('<p style="color:#94a3b8;font-size:14px;margin-bottom:20px">Entre com suas credenciais para acessar o painel</p>', unsafe_allow_html=True)
                with st.form("login_form"):
                    email = st.text_input("E-mail", placeholder="seu@email.com")
                    senha = st.text_input("Senha", type="password", placeholder="••••••••")
                    if st.form_submit_button("→  Acessar meu painel", use_container_width=True):
                        if not email or not senha:
                            st.warning("Preencha e-mail e senha.")
                        else:
                            res = requests.post(
                                f"{BASE_URL}/auth/login",
                                json={"email": email, "password": senha},
                                timeout=10,
                            )
                            if res.status_code == 200:
                                data  = res.json()
                                token = data.get("authToken")
                                # Limpa sessão anterior completamente
                                st.session_state.clear()
                                _invalidar_cache()
                                st.session_state.auth_token = token
                                st.session_state.logged_in  = True
                                try:
                                    me = requests.get(
                                        f"{BASE_URL}/auth/me",
                                        headers={"Authorization": f"Bearer {token}"},
                                        timeout=10,
                                    )
                                    if me.status_code == 200:
                                        me_data = me.json()
                                        st.session_state.user_name = me_data.get("name", me_data.get("email", email))
                                    else:
                                        st.session_state.user_name = email
                                except Exception:
                                    st.session_state.user_name = email
                                st.query_params['token'] = token
                                st.rerun()
                            else:
                                st.error("E-mail ou senha incorretos.")

            with tab_cadastro:
                st.markdown("### Crie sua conta")
                st.markdown('<p style="color:#94a3b8;font-size:14px;margin-bottom:20px">Preencha os dados abaixo para começar</p>', unsafe_allow_html=True)
                with st.form("cadastro_form"):
                    c1, c2 = st.columns(2)
                    with c1:
                        nome      = st.text_input("Nome", placeholder="Seu nome")
                    with c2:
                        sobrenome = st.text_input("Sobrenome", placeholder="Seu sobrenome (opcional)")
                    email_c = st.text_input("E-mail", placeholder="seu@email.com")
                    c3, c4  = st.columns(2)
                    with c3:
                        pass_c  = st.text_input("Senha", type="password", placeholder="Mínimo 8 caracteres")
                    with c4:
                        pass_c2 = st.text_input("Confirmar senha", type="password", placeholder="Repita a senha")
                    if st.form_submit_button("Criar minha conta", use_container_width=True):
                        nome_completo = f"{nome} {sobrenome}".strip()
                        if not nome_completo or not email_c or not pass_c:
                            st.warning("Preencha todos os campos obrigatórios.")
                        elif len(pass_c) < 8:
                            st.error("❌ A senha deve ter no mínimo 8 caracteres.")
                        elif pass_c != pass_c2:
                            st.error("❌ As senhas não coincidem.")
                        else:
                            payload = {"name": nome_completo, "email": email_c, "password": pass_c}
                            res = requests.post(f"{BASE_URL}/auth/signup", json=payload, timeout=10)
                            if res.status_code == 200:
                                # Login automático após cadastro
                                login_res = requests.post(
                                    f"{BASE_URL}/auth/login",
                                    json={"email": email_c, "password": pass_c},
                                    timeout=10,
                                )
                                if login_res.status_code == 200:
                                    login_data = login_res.json()
                                    # Limpa qualquer dado de sessão anterior
                                    token = login_data.get("authToken")
                                    st.session_state.clear()
                                    _invalidar_cache()
                                    st.session_state.auth_token = token
                                    st.session_state.logged_in  = True
                                    # Busca nome do novo usuário
                                    try:
                                        me = requests.get(
                                            f"{BASE_URL}/auth/me",
                                            headers={"Authorization": f"Bearer {token}"},
                                            timeout=10,
                                        )
                                        if me.status_code == 200:
                                            me_data = me.json()
                                            st.session_state.user_name = me_data.get("name", nome_completo)
                                        else:
                                            st.session_state.user_name = nome_completo
                                    except Exception:
                                        st.session_state.user_name = nome_completo
                                    st.query_params['token'] = token
                                    st.rerun()
                                else:
                                    st.success("Conta criada! Faça o login.")
                            else:
                                try:
                                    err_msg = res.json().get('message', '')
                                    if 'email' in err_msg.lower() or 'unique' in err_msg.lower():
                                        st.error("❌ Este e-mail já está cadastrado. Tente fazer login.")
                                    else:
                                        st.error(f"❌ Erro ao cadastrar: {err_msg or 'Tente novamente.'}")
                                except Exception:
                                    st.error("❌ Erro ao cadastrar. Tente novamente.")


# ------------------------------------------------
# 6. MÓDULOS CRUD
# ------------------------------------------------

# GESTÃO DE PROFESSORES

def modulo_professores():
    '''
    Gerenciamento de professores com cards visuais.
    Cada card exibe: avatar com iniciais, nome, e-mail,
    disciplinas vinculadas com média de notas e média geral.
    Ao remover um professor, aplica deleção em cascata:
    tarefas → metas → anotações → disciplinas → professor.
    '''
    st.header('👨‍🏫 Meus Professores')
    st.caption('Cadastre e gerencie seus professores')

    # Busca dados
    dados   = api_get('professores', st.session_state.get('auth_token', '')) or []
    discs   = api_get('disciplinas', st.session_state.get('auth_token', '')) or []
    tarefas = api_get('tarefas', st.session_state.get('auth_token', ''))     or []

    # Mapa prof_id → disciplinas
    disc_por_prof = {}
    for d in discs:
        pid = d.get('prof_id')
        if pid:
            disc_por_prof.setdefault(pid, []).append(d)

    # Mapa disc_id → média das notas
    df_t = pd.DataFrame(tarefas) if tarefas else pd.DataFrame()
    media_por_disc = {}
    if not df_t.empty and 'nota' in df_t.columns and 'disc_id' in df_t.columns:
        media_por_disc = df_t.groupby('disc_id')['nota'].mean().to_dict()

    # ── Formulário adicionar ──────────────────────────────
    with st.expander('➕  Adicionar novo professor'):
        with st.form('form_add_prof'):
            c1, c2 = st.columns(2)
            with c1:
                nome  = st.text_input('Nome do Professor', placeholder='Ex: João Silva')
            with c2:
                email = st.text_input('E-mail de Contato', placeholder='joao@escola.com')
            if st.form_submit_button('Cadastrar Professor', type='primary'):
                if nome and email:
                    res = api_post('professores', {'nome': nome, 'email': email})
                    if res and res.status_code in (200, 201):
                        st.success('Professor cadastrado!')
                        _invalidar_cache()
                        st.rerun()
                    else:
                        st.error('Erro ao cadastrar professor.')
                else:
                    st.warning('Preencha nome e e-mail.')

    if not dados:
        st.info('Nenhum professor cadastrado ainda.')
        return

    st.markdown(f'**{len(dados)} professor(es) cadastrado(s)**')
    st.markdown('---')

    # ── Cards em grid ─────────────────────────────────────
    CORES_AV = [
        ('#EEEDFE', '#3C3489'), ('#E1F5EE', '#085041'),
        ('#FAECE7', '#712B13'), ('#E6F1FB', '#0C447C'),
        ('#FAEEDA', '#633806'), ('#FBEAF0', '#72243E'),
    ]

    cols = st.columns(3)
    for i, prof in enumerate(dados):
        prof_id   = prof['id']
        prof_nome = prof.get('nome', '')
        prof_mail = prof.get('email', '')
        iniciais  = ''.join([p[0].upper() for p in prof_nome.split()[:2]]) if prof_nome else '??'
        bg_av, tc_av = CORES_AV[i % len(CORES_AV)]

        # Disciplinas deste professor e médias
        discs_prof = disc_por_prof.get(prof_id, [])
        medias_discs = []
        for d in discs_prof:
            m = media_por_disc.get(d['id'])
            if m is not None:
                medias_discs.append(float(m))

        media_geral_prof = sum(medias_discs) / len(medias_discs) if medias_discs else None

        with cols[i % 3]:
            media_txt = f'{media_geral_prof:.1f}' if media_geral_prof is not None else '—'

            # Monta disciplinas HTML
            discs_html = ''
            if discs_prof:
                for d in discs_prof:
                    m = media_por_disc.get(d['id'])
                    if m is not None:
                        m = float(m)
                        cor_m = '#166534' if m >= 7 else '#92400e' if m >= 5 else '#991b1b'
                        bg_m  = '#dcfce7' if m >= 7 else '#fef9e7' if m >= 5 else '#fee2e2'
                        nota_span = f'<span style="font-size:11px;font-weight:500;color:{cor_m};background:{bg_m};padding:2px 7px;border-radius:4px">{m:.1f}</span>'
                    else:
                        nota_span = '<span style="font-size:11px;color:#94a3b8;font-style:italic">sem notas</span>'
                    discs_html += (
                        f'<div style="display:flex;align-items:center;justify-content:space-between;'
                        f'padding:5px 8px;border-radius:6px;background:rgba(128,128,128,0.08);'
                        f'margin-bottom:4px">'
                        f'<span style="font-size:11px;color:var(--color-text-secondary);'
                        f'white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:130px">'
                        f'{d["nome"][:24]}</span>{nota_span}</div>'
                    )
            else:
                discs_html = '<div style="font-size:12px;color:var(--color-text-secondary);font-style:italic;padding:4px 0">Sem disciplinas vinculadas</div>'

            st.markdown(f'''
            <div style="background:var(--color-background-primary);
                        border:0.5px solid var(--color-border-tertiary);
                        border-radius:12px;padding:16px;margin-bottom:4px">
              <div style="display:flex;align-items:center;gap:10px;margin-bottom:12px">
                <div style="width:42px;height:42px;border-radius:50%;background:{bg_av};
                  display:flex;align-items:center;justify-content:center;
                  font-size:14px;font-weight:500;color:{tc_av};flex-shrink:0">{iniciais}</div>
                <div style="min-width:0;flex:1">
                  <div style="font-size:14px;font-weight:500;color:var(--color-text-primary)">{prof_nome}</div>
                  <div style="font-size:11px;color:var(--color-text-secondary);
                    white-space:nowrap;overflow:hidden;text-overflow:ellipsis">{prof_mail}</div>
                </div>
              </div>
              <div style="margin-bottom:10px">{discs_html}</div>
              <div style="border-top:0.5px solid var(--color-border-tertiary);padding-top:10px;
                display:flex;align-items:center;justify-content:space-between">
                <div>
                  <div style="font-size:10px;color:var(--color-text-secondary);
                    text-transform:uppercase;letter-spacing:.05em;margin-bottom:2px">Média geral</div>
                  <div style="font-size:20px;font-weight:500;color:var(--color-text-primary)">{media_txt}</div>
                </div>
              </div>
            </div>
            ''', unsafe_allow_html=True)

            # Botões editar/remover
            col_e, col_r, _ = st.columns([1, 1, 2])
            with col_e:
                if st.button('✏️', key=f'edit_prof_{prof_id}', help='Editar'):
                    st.session_state['editando_prof']  = prof_id
                    st.session_state.pop(f'confirm_del_{prof_id}', None)
                    st.rerun()
            with col_r:
                if st.button('🗑️', key=f'del_prof_{prof_id}', help='Remover'):
                    st.session_state[f'confirm_del_{prof_id}'] = True

            # Confirmação de remoção
            if st.session_state.get(f'confirm_del_{prof_id}'):
                st.warning(f'Remover **{prof_nome}**? Essa ação não pode ser desfeita.')
                col_sim, col_nao, _ = st.columns([1, 1, 3])
                with col_sim:
                    if st.button('✅ Sim', key=f'sim_del_{prof_id}', type='primary'):
                        # ── Deleção em cascata ─────────────────────────
                        discs_do_prof = disc_por_prof.get(prof_id, [])
                        ids_discs_del = {d['id'] for d in discs_do_prof}

                        if ids_discs_del:
                            token         = st.session_state.get('auth_token', '')
                            tarefas_all   = api_get('tarefas',   token) or []
                            metas_all     = api_get('metas',     token) or []
                            anots_all     = api_get('anotacoes', token) or []

                            # 1. Deleta tarefas das disciplinas
                            for t in tarefas_all:
                                if t.get('disc_id') in ids_discs_del:
                                    api_delete('tarefas', t['id'])

                            # 2. Deleta metas das disciplinas
                            for m in metas_all:
                                if m.get('disc_id') in ids_discs_del:
                                    api_delete('metas', m['id'])

                            # 3. Deleta anotações das disciplinas
                            for a in anots_all:
                                if a.get('disc_id') in ids_discs_del:
                                    api_delete('anotacoes', a['id'])

                            # 4. Deleta as disciplinas
                            for d in discs_do_prof:
                                api_delete('disciplinas', d['id'])

                        # 5. Deleta o professor
                        res = api_delete('professores', prof_id)
                        if res and res.status_code in (200, 204):
                            st.session_state.pop(f'confirm_del_{prof_id}', None)
                            _invalidar_cache()
                            st.rerun()
                        else:
                            status = res.status_code if res else 'sem resposta'
                            msg    = res.text[:150] if res else ''
                            st.error(f'Erro ao remover professor. Status: {status} — {msg}')
                with col_nao:
                    if st.button('❌ Não', key=f'nao_del_{prof_id}'):
                        st.session_state.pop(f'confirm_del_{prof_id}', None)
                        st.rerun()

            st.markdown('---')

    # ── Modal de edição ───────────────────────────────────
    prof_edit_id = st.session_state.get('editando_prof')
    if prof_edit_id:
        prof_edit = next((p for p in dados if p['id'] == prof_edit_id), None)
        if prof_edit:
            st.subheader(f'✏️ Editando: {prof_edit.get("nome", "")}')
            with st.form('form_edit_prof'):
                c1, c2 = st.columns(2)
                with c1:
                    novo_nome  = st.text_input('Nome', value=prof_edit.get('nome', ''))
                with c2:
                    novo_email = st.text_input('E-mail', value=prof_edit.get('email', ''))
                col_s, col_c, _ = st.columns([1, 1, 3])
                with col_s:
                    if st.form_submit_button('💾 Salvar', type='primary'):
                        res = api_patch('professores', prof_edit_id,
                                        {'nome': novo_nome, 'email': novo_email})
                        if res and res.status_code in (200, 201, 204):
                            st.success('Professor atualizado!')
                            st.session_state.pop('editando_prof', None)
                            _invalidar_cache()
                            st.rerun()
                        else:
                            st.error('Erro ao atualizar.')
                with col_c:
                    if st.form_submit_button('Cancelar'):
                        st.session_state.pop('editando_prof', None)
                        st.rerun()


# GESTÃO DE DISCIPLINAS

def modulo_disciplinas():
    col_titulo, col_btn = st.columns([4, 1])
    with col_titulo:
        st.header('📖 Minhas Disciplinas')
        st.caption('Gerencie as matérias do seu curso')

    profs = api_get('professores', st.session_state.get('auth_token', ''))
    if not profs:
        st.markdown("""
        <div style="text-align:center;padding:60px 20px">
          <div style="font-size:52px;margin-bottom:16px">📖</div>
          <div style="font-size:22px;font-weight:500;color:#f1f5f9;margin-bottom:10px">
            Nenhum professor cadastrado
          </div>
          <div style="font-size:15px;color:#94a3b8;margin-bottom:32px;line-height:1.7">
            Para criar uma disciplina você precisa primeiro<br>ter um professor cadastrado.
          </div>
        </div>
        """, unsafe_allow_html=True)
        col_a, col_b, col_c = st.columns([2, 1, 2])
        with col_b:
            if st.button('+ Cadastrar professor', key='disc_goto_prof', use_container_width=True):
                st.session_state['_nav_target'] = 'Professores'
                st.rerun()
        return

    with st.expander('➕  Nova Disciplina'):
        with st.form('form_add_disc'):
            c1, c2 = st.columns(2)
            with c1:
                nome_d = st.text_input('Nome da Matéria', placeholder='Ex: Matemática')
            with c2:
                opcoes_p   = {p['nome']: p['id'] for p in profs}
                p_escolhido = st.selectbox('Professor Responsável', options=list(opcoes_p.keys()))
            if st.form_submit_button('Salvar Disciplina', type='primary'):
                if nome_d:
                    res = api_post('disciplinas',
                                   {'nome': nome_d, 'prof_id': opcoes_p[p_escolhido]})
                    if res and res.status_code in (200, 201):
                        st.success('Disciplina criada!')
                        _invalidar_cache()
                        st.rerun()
                    else:
                        st.error('Erro ao criar disciplina.')
                else:
                    st.warning('Digite o nome da disciplina.')

    discs = api_get('disciplinas', st.session_state.get('auth_token', ''))
    if discs:
        df_d   = pd.DataFrame(discs)
        df_p   = pd.DataFrame(profs)
        df_view = df_d.merge(df_p[['id', 'nome']],
                             left_on='prof_id', right_on='id',
                             suffixes=('', '_prof'))
        st.dataframe(
            df_view[['id', 'nome', 'nome_prof']].rename(
                columns={'nome': 'Disciplina', 'nome_prof': 'Professor'}
            ),
            use_container_width=True,
            hide_index=True,
        )

        st.markdown('---')

        # FEATURE: Calculadora de média necessária
        st.subheader('🧮 Calculadora de Média')
        st.caption('Descubra qual nota precisa tirar para atingir a média')
        tarefas_calc = api_get('tarefas', st.session_state.get('auth_token', ''))
        if tarefas_calc:
            df_calc = pd.DataFrame(tarefas_calc)
            if 'nota' in df_calc.columns and len(discs) > 0:
                c_calc1, c_calc2, c_calc3 = st.columns(3)
                with c_calc1:
                    disc_calc = st.selectbox(
                        'Disciplina',
                        options=[d['nome'] for d in discs],
                        key='calc_disc',
                    )
                with c_calc2:
                    media_alvo = st.number_input('Média desejada', 0.0, 10.0, 7.0, step=0.5, key='calc_media')
                with c_calc3:
                    total_ativs = st.number_input('Total de atividades previstas', 1, 20, 4, step=1, key='calc_total')

                # calcula a média atual nessa disciplina
                disc_id_calc = next((d['id'] for d in discs if d['nome'] == disc_calc), None)
                if disc_id_calc:
                    notas_disc = df_calc[df_calc['disc_id'] == disc_id_calc]['nota'].dropna().tolist()
                    qtd_atual  = len(notas_disc)
                    media_atual = sum(notas_disc) / qtd_atual if qtd_atual > 0 else 0
                    restantes   = total_ativs - qtd_atual

                    if restantes <= 0:
                        nota_necessaria = None
                    else:
                        # média_alvo = (soma_atual + nota_necessaria * restantes) / total
                        nota_necessaria = (media_alvo * total_ativs - sum(notas_disc)) / restantes

                    col_r1, col_r2, col_r3 = st.columns(3)
                    col_r1.metric('Média atual', f'{media_atual:.1f}')
                    col_r2.metric('Atividades feitas', f'{qtd_atual}/{total_ativs}')
                    if restantes <= 0:
                        col_r3.metric('Situação', '✅ Completo')
                    elif nota_necessaria is not None and nota_necessaria <= 0:
                        col_r3.metric('Nota necessária', '✅ Já atingiu!')
                    elif nota_necessaria is not None and nota_necessaria > 10:
                        col_r3.metric('Nota necessária', '❌ Impossível', delta=f'{nota_necessaria:.1f}')
                    elif nota_necessaria is not None:
                        col_r3.metric('Nota necessária', f'{nota_necessaria:.1f}', delta=f'{restantes} ativ. restantes')

        st.markdown('---')
        st.subheader('🗑️ Remover Disciplina')
        id_del = st.selectbox(
            'Selecione a disciplina para remover',
            options=[d['id'] for d in discs],
            format_func=lambda x: next(
                (d['nome'] for d in discs if d['id'] == x), str(x)
            ),
        )
        if st.button('Remover Disciplina', type='primary'):
            token       = st.session_state.get('auth_token', '')
            tarefas_all = api_get('tarefas',   token) or []
            metas_all   = api_get('metas',     token) or []
            anots_all   = api_get('anotacoes', token) or []

            for t in tarefas_all:
                if t.get('disc_id') == id_del:
                    api_delete('tarefas', t['id'])
            for m in metas_all:
                if m.get('disc_id') == id_del:
                    api_delete('metas', m['id'])
            for a in anots_all:
                if a.get('disc_id') == id_del:
                    api_delete('anotacoes', a['id'])

            res = api_delete('disciplinas', id_del)
            if res and res.status_code in (200, 204):
                st.success('Disciplina e dados relacionados removidos.')
                _invalidar_cache()
                st.rerun()
            else:
                st.error('Erro ao remover disciplina.')


# GESTÃO DE TAREFAS

def modulo_tarefas():
    '''
    Quadro de atividades e notas com funcionalidades avançadas:
      - Filtros por disciplina, status, prioridade + busca por nome
      - Ordenação por prazo, nota, prioridade ou nome
      - Checkbox de conclusão que salva no Xano via PATCH
      - Exportação para Excel com aba Dashboard de métricas
      - Edição inline de nome, nota, prazo e prioridade
      - Contador de prazo com alertas visuais (Hoje!, Atrasado, Xd)
    '''
    st.header('✅ Minhas Tarefas e Notas')
    st.caption('Registre atividades e acompanhe seu desempenho')

    discs = api_get('disciplinas', st.session_state.get('auth_token', ''))
    if not discs:
        st.markdown("""
        <div style="text-align:center;padding:60px 20px">
          <div style="font-size:52px;margin-bottom:16px">✅</div>
          <div style="font-size:22px;font-weight:500;color:#f1f5f9;margin-bottom:10px">
            Nenhuma disciplina cadastrada
          </div>
          <div style="font-size:15px;color:#94a3b8;margin-bottom:32px;line-height:1.7">
            Suas atividades e notas ficam organizadas por disciplina.<br>
            Cadastre uma disciplina para começar a lançar notas.
          </div>
        </div>
        """, unsafe_allow_html=True)
        col_a, col_b, col_c = st.columns([2, 1, 2])
        with col_b:
            if st.button('+ Cadastrar disciplina', key='tar_goto_disc', use_container_width=True):
                st.session_state['_nav_target'] = 'Disciplinas'
                st.rerun()
        return

    with st.expander('➕  Lançar nova Atividade / Nota'):
        with st.form('form_add_tarefa'):
            c1, c2, c3, c4 = st.columns([3, 2, 1, 1])
            with c1:
                nome_t = st.text_input('Nome da Atividade', placeholder='Ex: Prova 1 — Álgebra')
            with c2:
                opcoes_d    = {d['nome']: d['id'] for d in discs}
                d_escolhida = st.selectbox('Disciplina', options=list(opcoes_d.keys()))
            with c3:
                nota = st.number_input('Nota', 0.0, 10.0, 0.0, step=0.1)
            with c4:
                prioridade = st.selectbox('Prioridade', ['🔴 Alta', '🟡 Média', '🟢 Baixa'])
            c5, c6 = st.columns(2)
            with c5:
                concluida = st.checkbox('Marcar como concluída')
            with c6:
                from datetime import datetime as _dt
                prazo = st.date_input('Prazo de entrega', value=None)
            if st.form_submit_button('Registrar Atividade', type='primary'):
                if nome_t:
                    res = api_post('tarefas', {
                        'nome':       nome_t,
                        'disc_id':    opcoes_d[d_escolhida],
                        'nota':       float(nota),
                        'concluida':  bool(concluida),
                        'prioridade': prioridade,
                        'prazo':      prazo.isoformat() if prazo else None,
                    })
                    if res and res.status_code in (200, 201):
                        st.success('Atividade registrada!')
                        _invalidar_cache()
                        st.rerun()
                    else:
                        st.error('Erro ao registrar atividade.')
                else:
                    st.warning('Digite o nome da atividade.')

    tarefas = api_get('tarefas', st.session_state.get('auth_token', ''))
    if not tarefas:
        st.info('Nenhuma atividade cadastrada ainda.')
        return

    df_t = pd.DataFrame(tarefas)
    if 'concluida' not in df_t.columns:
        df_t['concluida'] = False
    df_t['concluida'] = df_t['concluida'].fillna(False).astype(bool)

    nomes_disc = {d['id']: d['nome'] for d in discs}
    df_t['disciplina'] = df_t['disc_id'].map(nomes_disc).fillna('—')

    # ── Busca + Filtros ───────────────────────────────────
    col_busca, _ = st.columns([2, 3])
    with col_busca:
        busca = st.text_input('🔍 Buscar atividade', placeholder='Digite o nome da atividade...',
                              label_visibility='collapsed')

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        opcoes_disc = list(nomes_disc.values())
        filtro_disc = st.multiselect('Disciplina', options=opcoes_disc, default=opcoes_disc)
    with col2:
        filtro_status = st.multiselect('Status', options=['Concluída', 'Pendente'],
                                       default=['Concluída', 'Pendente'])
    with col3:
        prios_disp  = ['Todas', '🔴 Alta', '🟡 Média', '🟢 Baixa']
        filtro_prio = st.selectbox('Prioridade', options=prios_disp)
    with col4:
        ordenar_por = st.selectbox('Ordenar por',
            options=['Padrão', '📅 Prazo', '📉 Menor nota', '📈 Maior nota', '🔴 Prioridade', '🔤 Nome'])

    df_filtrado = df_t[df_t['disciplina'].isin(filtro_disc)].copy()
    status_mask = pd.Series([False] * len(df_filtrado), index=df_filtrado.index)
    if 'Concluída' in filtro_status:
        status_mask |= df_filtrado['concluida'] == True
    if 'Pendente' in filtro_status:
        status_mask |= df_filtrado['concluida'] == False
    df_filtrado = df_filtrado[status_mask]

    if 'prioridade' in df_filtrado.columns and filtro_prio != 'Todas':
        df_filtrado = df_filtrado[df_filtrado['prioridade'] == filtro_prio]

    # Aplica busca por nome
    if busca.strip():
        df_filtrado = df_filtrado[
            df_filtrado['nome'].str.contains(busca.strip(), case=False, na=False)
        ]

    # Aplica ordenação
    if not df_filtrado.empty and 'ordenar_por' in dir():
        if ordenar_por == '📅 Prazo (mais próximo)' and 'prazo' in df_filtrado.columns:
            df_filtrado = df_filtrado.copy()
            df_filtrado['_prazo_sort'] = pd.to_datetime(
                df_filtrado['prazo'].replace('None', pd.NaT), errors='coerce'
            )
            df_filtrado = df_filtrado.sort_values('_prazo_sort', na_position='last')
            df_filtrado = df_filtrado.drop(columns=['_prazo_sort'])
        elif ordenar_por == '📉 Nota (menor primeiro)' and 'nota' in df_filtrado.columns:
            df_filtrado = df_filtrado.sort_values('nota', ascending=True)
        elif ordenar_por == '📈 Nota (maior primeiro)' and 'nota' in df_filtrado.columns:
            df_filtrado = df_filtrado.sort_values('nota', ascending=False)
        elif ordenar_por == '🔴 Prioridade' and 'prioridade' in df_filtrado.columns:
            ordem_prio  = {'🔴 Alta': 0, '🟡 Média': 1, '🟢 Baixa': 2}
            df_filtrado = df_filtrado.copy()
            df_filtrado['_prio_ord'] = df_filtrado['prioridade'].map(ordem_prio).fillna(1)
            df_filtrado = df_filtrado.sort_values('_prio_ord').drop(columns=['_prio_ord'])
        elif ordenar_por == '🔤 Nome' and 'nome' in df_filtrado.columns:
            df_filtrado = df_filtrado.sort_values('nome')

    if df_filtrado.empty:
        st.info('Nenhuma atividade encontrada com os filtros selecionados.')
    else:
        from datetime import datetime as _dt2

        # FEATURE: calcula dias restantes para o prazo
        if 'prazo' not in df_filtrado.columns:
            df_filtrado['prazo'] = None
        if 'prioridade' not in df_filtrado.columns:
            df_filtrado['prioridade'] = '🟡 Média'

        def calcular_prazo(prazo_str):
            '''Retorna string de dias restantes ou vazio se sem prazo.'''
            if not prazo_str or str(prazo_str) in ('None', 'nan', ''):
                return '—'
            try:
                prazo_dt = _dt2.fromisoformat(str(prazo_str)[:10])
                dias = (prazo_dt.date() - _dt2.today().date()).days
                if dias < 0:
                    return '⚠️ Atrasado'
                elif dias == 0:
                    return '🔴 Hoje!'
                elif dias <= 3:
                    return f'🟡 {dias}d'
                else:
                    return f'🟢 {dias}d'
            except Exception:
                return '—'

        df_filtrado = df_filtrado.copy()
        df_filtrado['dias_restantes'] = df_filtrado['prazo'].apply(calcular_prazo)

        # Monta colunas — id oculto mas mantido para mapeamento interno
        colunas_exib = ['id', 'nome', 'disciplina', 'nota', 'concluida']
        rename_map = {
            'nome':       'Atividade',
            'disciplina': 'Disciplina',
            'nota':       'Nota',
            'concluida':  'Concluída',
        }
        col_config = {
            'id':         st.column_config.Column('id', disabled=True),
            'Concluída':  st.column_config.CheckboxColumn('✅'),
            'Nota':       st.column_config.NumberColumn('Nota', format='%.1f'),
            'Atividade':  st.column_config.TextColumn('Atividade',  width='large'),
            'Disciplina': st.column_config.TextColumn('Disciplina', width='medium'),
        }

        if 'prioridade' in df_filtrado.columns:
            colunas_exib.append('prioridade')
            rename_map['prioridade'] = 'Prioridade'
            col_config['Prioridade'] = st.column_config.TextColumn('Prioridade', width='small')

        colunas_exib.append('dias_restantes')
        rename_map['dias_restantes'] = 'Prazo'
        col_config['Prazo'] = st.column_config.TextColumn('Prazo', width='small')

        disabled_cols = ['id', 'Atividade', 'Disciplina', 'Nota']
        if 'Prioridade' in rename_map.values() and 'prioridade' in df_filtrado.columns:
            disabled_cols.append('Prioridade')
        disabled_cols.append('Prazo')

        df_editado = st.data_editor(
            df_filtrado[colunas_exib].rename(columns=rename_map),
            use_container_width=True,
            hide_index=True,
            disabled=disabled_cols,
            column_config=col_config,
            column_order=['Atividade', 'Disciplina', 'Nota', 'Prioridade', 'Prazo', '✅', 'id'],
            key='editor_tarefas',
        )

        col_btn1, col_btn2 = st.columns([1, 1])
        with col_btn1:
            if st.button('💾 Salvar', type='primary'):
                # CORREÇÃO DEFINITIVA:
                # Quando o botão é clicado o Streamlit reroda o script inteiro
                # e df_editado perde as edições do usuário.
                # A solução é ler st.session_state['editor_tarefas'] que guarda
                # exatamente quais linhas foram editadas e quais valores mudaram.
                # Formato: {'edited_rows': {idx: {'Concluída': True/False}, ...}}

                estado_editor = st.session_state.get('editor_tarefas', {})
                linhas_editadas = estado_editor.get('edited_rows', {})

                if not linhas_editadas:
                    st.info('Nenhuma alteração detectada. Marque ou desmarque um checkbox antes de salvar.')
                else:
                    # Monta lista ordenada do df exibido para mapear índice → id real
                    df_exibido = df_filtrado[colunas_exib].rename(columns=rename_map).reset_index(drop=True)

                    alteracoes = 0
                    erros      = 0

                    for idx_str, campos in linhas_editadas.items():
                        if 'Concluída' not in campos:
                            continue  # só processa mudanças no checkbox

                        try:
                            idx       = int(idx_str)
                            tarefa_id = int(df_exibido.loc[idx, 'id'])
                            nova      = bool(campos['Concluída'])
                            res = api_patch('tarefas', tarefa_id, {'concluida': bool(nova)})
                            if res is None:
                                erros += 1
                            elif res.status_code in (200, 201, 204):
                                alteracoes += 1
                            else:
                                st.warning(f'Tarefa {tarefa_id}: status {res.status_code} — {res.text[:120]}')
                                erros += 1
                        except Exception as e:
                            st.warning(f'Erro ao processar linha {idx_str}: {e}')
                            erros += 1

                    _invalidar_cache()
                    if alteracoes > 0:
                        st.success(f'✅ {alteracoes} tarefa(s) atualizada(s)!')
                    if erros > 0:
                        st.error(f'❌ {erros} tarefa(s) com erro ao salvar.')
                st.rerun()

        with col_btn2:
            import io as _io
            import openpyxl
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                          Border, Side, GradientFill)
            from openpyxl.utils import get_column_letter

            buf_excel = _io.BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Notas'

            # Cores
            COR_HEADER  = 'FF0F172A'  # escuro
            COR_VERDE   = 'FFDCFCE7'
            COR_AMARELO = 'FFFEF9E7'
            COR_VERM    = 'FFFEE2E2'
            COR_ALT     = 'FFF8FAFC'

            bd = Side(style='thin', color='FFE2E8F0')
            borda = Border(left=bd, right=bd, top=bd, bottom=bd)

            # Linha de título
            ws.merge_cells('A1:F1')
            ws['A1'] = f'EduTrack AI — Notas e Atividades  ·  {date.today().strftime("%d/%m/%Y")}'
            ws['A1'].font      = Font(bold=True, size=13, color='FF0F172A')
            ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[1].height = 28

            # Cabeçalhos
            headers = ['Atividade', 'Disciplina', 'Nota', 'Prioridade', 'Prazo', 'Status']
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=col_idx, value=h)
                cell.font      = Font(bold=True, color='FFFFFFFF', size=10)
                cell.fill      = PatternFill('solid', fgColor=COR_HEADER)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border    = borda
            ws.row_dimensions[2].height = 22

            # Dados
            cols_export = ['nome', 'disciplina', 'nota']
            if 'prioridade'    in df_filtrado.columns: cols_export.append('prioridade')
            else: cols_export.append('nome')
            cols_export.append('dias_restantes')
            cols_export.append('concluida')

            for row_idx, (_, row) in enumerate(df_filtrado.iterrows(), start=3):
                nota_v   = float(row.get('nota', 0) or 0)
                conc_v   = bool(row.get('concluida', False))
                status_v = 'Concluída' if conc_v else 'Pendente'

                cor_linha = COR_VERDE   if nota_v >= 7 else                             COR_AMARELO if nota_v >= 5 else                             COR_VERM    if nota_v > 0  else                             COR_ALT

                valores = [
                    str(row.get('nome', '')),
                    str(row.get('disciplina', '')),
                    nota_v,
                    str(row.get('prioridade', '') or ''),
                    str(row.get('dias_restantes', '') or '—'),
                    status_v,
                ]
                for col_idx, val in enumerate(valores, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.fill      = PatternFill('solid', fgColor=cor_linha)
                    cell.border    = borda
                    cell.alignment = Alignment(vertical='center',
                                               horizontal='center' if col_idx in (3,4,5,6) else 'left')
                    if col_idx == 3:
                        cell.font = Font(bold=True, size=10)
                ws.row_dimensions[row_idx].height = 18

            # Larguras das colunas
            ws.column_dimensions['A'].width = 38
            ws.column_dimensions['B'].width = 28
            ws.column_dimensions['C'].width = 8
            ws.column_dimensions['D'].width = 14
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12

            # Congela cabeçalho
            ws.freeze_panes = 'A3'

            # ── ABA DASHBOARD ────────────────────────────────
            from openpyxl.chart import BarChart, Reference
            from openpyxl.chart.series import DataPoint

            wd = wb.create_sheet(title='Dashboard')
            wd.sheet_view.showGridLines = False

            def cel(ws, row, col, value='', bold=False, size=11,
                    color='FF0F172A', bg=None, align='left', border=False):
                c = ws.cell(row=row, column=col, value=value)
                c.font      = Font(bold=bold, size=size, color=color)
                c.alignment = Alignment(horizontal=align, vertical='center',
                                        wrap_text=True)
                if bg:
                    c.fill = PatternFill('solid', fgColor=bg)
                if border:
                    bd2 = Side(style='thin', color='FFE2E8F0')
                    c.border = Border(left=bd2, right=bd2, top=bd2, bottom=bd2)
                return c

            # Título
            wd.merge_cells('B2:H2')
            cel(wd, 2, 2, 'EduTrack AI — Dashboard de Desempenho',
                bold=True, size=16, color='FF0F172A')
            wd.merge_cells('B3:H3')
            cel(wd, 3, 2, f'Gerado em {date.today().strftime("%d/%m/%Y")}',
                size=10, color='FF64748B')
            wd.row_dimensions[2].height = 30
            wd.row_dimensions[3].height = 16

            # Métricas
            total_t  = len(df_filtrado)
            conc_t   = int(df_filtrado['concluida'].sum()) if 'concluida' in df_filtrado.columns else 0
            pend_t   = total_t - conc_t
            media_t  = float(df_filtrado['nota'].mean()) if 'nota' in df_filtrado.columns and total_t > 0 else 0.0
            melhor_t = float(df_filtrado['nota'].max())  if 'nota' in df_filtrado.columns and total_t > 0 else 0.0
            pior_t   = float(df_filtrado['nota'].min())  if 'nota' in df_filtrado.columns and total_t > 0 else 0.0

            metricas_cards = [
                ('Atividades',    total_t,          'FF0F172A', 'FFF1F5F9'),
                ('Concluídas',    conc_t,            'FF166534', 'FFDCFCE7'),
                ('Pendentes',     pend_t,            'FF92400E', 'FFFEF9E7'),
                ('Média Geral',   f'{media_t:.1f}',  'FF92400E', 'FFFFF5C8'),
                ('Melhor Nota',   f'{melhor_t:.1f}', 'FF166534', 'FFDCFCE7'),
                ('Pior Nota',     f'{pior_t:.1f}',   'FF991B1B', 'FFFEE2E2'),
            ]

            col_start = 2
            for i, (label, valor, tc, bg) in enumerate(metricas_cards):
                col = col_start + i
                wd.merge_cells(
                    start_row=5, start_column=col,
                    end_row=5,   end_column=col
                )
                wd.merge_cells(
                    start_row=6, start_column=col,
                    end_row=6,   end_column=col
                )
                cel(wd, 5, col, str(valor), bold=True, size=20,
                    color=tc, bg=bg, align='center', border=True)
                cel(wd, 6, col, label, bold=False, size=9,
                    color='FF64748B', bg=bg, align='center', border=True)
                wd.row_dimensions[5].height = 36
                wd.row_dimensions[6].height = 20
                wd.column_dimensions[get_column_letter(col)].width = 14

            # Título tabela por disciplina
            wd.merge_cells('B9:H9')
            cel(wd, 9, 2, 'Desempenho por Disciplina',
                bold=True, size=12, color='FF0F172A')
            wd.row_dimensions[9].height = 22

            # Cabeçalho
            disc_headers = ['Disciplina', 'Atividades', 'Concluídas', 'Pendentes', 'Média', 'Melhor', 'Pior']
            for ci, h in enumerate(disc_headers, start=2):
                cel(wd, 10, ci, h, bold=True, size=9,
                    color='FFFFFFFF', bg='FF0F172A', align='center', border=True)
            wd.row_dimensions[10].height = 20

            # Dados por disciplina
            discs_unicas = df_filtrado['disciplina'].unique() if 'disciplina' in df_filtrado.columns else []
            row_d = 11
            for disc_nome in discs_unicas:
                df_d = df_filtrado[df_filtrado['disciplina'] == disc_nome]
                total_d  = len(df_d)
                conc_d   = int(df_d['concluida'].sum()) if 'concluida' in df_d.columns else 0
                pend_d   = total_d - conc_d
                media_d  = float(df_d['nota'].mean()) if 'nota' in df_d.columns else 0.0
                melhor_d = float(df_d['nota'].max())  if 'nota' in df_d.columns else 0.0
                pior_d   = float(df_d['nota'].min())  if 'nota' in df_d.columns else 0.0

                cor_media = 'FFDCFCE7' if media_d >= 7 else 'FFFEF9E7' if media_d >= 5 else 'FFFEE2E2'
                bg_row    = 'FFF8FAFC' if row_d % 2 == 0 else 'FFFFFFFF'

                valores_d = [disc_nome, total_d, conc_d, pend_d,
                             f'{media_d:.1f}', f'{melhor_d:.1f}', f'{pior_d:.1f}']
                for ci, val in enumerate(valores_d, start=2):
                    bg_usar = cor_media if ci == 6 else bg_row
                    cel(wd, row_d, ci, val, size=9,
                        color='FF0F172A', bg=bg_usar,
                        align='left' if ci == 2 else 'center', border=True)
                wd.row_dimensions[row_d].height = 18
                row_d += 1

            # Larguras dashboard
            wd.column_dimensions['A'].width = 2
            wd.column_dimensions['B'].width = 28
            for col_l in ['C', 'D', 'E', 'F', 'G', 'H']:
                wd.column_dimensions[col_l].width = 13

            # Gráfico de barras da média por disciplina
            if len(discs_unicas) > 0:
                chart = BarChart()
                chart.type    = 'col'
                chart.title   = 'Média por Disciplina'
                chart.y_axis.title = 'Nota'
                chart.x_axis.title = 'Disciplina'
                chart.style   = 10
                chart.width   = 15
                chart.height  = 10

                data_ref  = Reference(wd, min_col=6, min_row=10,
                                      max_row=10 + len(discs_unicas))
                cats_ref  = Reference(wd, min_col=2, min_row=11,
                                      max_row=10 + len(discs_unicas))
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                chart.series[0].graphicalProperties.solidFill = 'F5C842'
                chart.series[0].graphicalProperties.line.solidFill = '92400E'

                wd.add_chart(chart, f'B{row_d + 2}')

            # Ordena abas — Dashboard primeiro
            wb.move_sheet('Dashboard', offset=-1)

            wb.save(buf_excel)
            buf_excel.seek(0)
            st.download_button(
                label='📊 Exportar Excel',
                data=buf_excel,
                file_name=f"notas_edutrack_{date.today().isoformat()}.xlsx",
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )

    st.markdown('---')

    tab_editar, tab_remover = st.tabs(['✏️ Editar Atividade', '🗑️ Remover Atividade'])

    todas_tarefas = api_get('tarefas', st.session_state.get('auth_token', ''))

    with tab_editar:
        if not todas_tarefas:
            st.info('Nenhuma atividade cadastrada ainda.')
        else:
            id_edit = st.selectbox(
                'Selecione a atividade para editar',
                options=[t['id'] for t in todas_tarefas],
                format_func=lambda x: next(
                    (t['nome'] for t in todas_tarefas if t['id'] == x), str(x)
                ),
                key='sel_editar_tarefa',
            )
            tarefa_sel = next((t for t in todas_tarefas if t['id'] == id_edit), None)

            if tarefa_sel:
                with st.form('form_editar_tarefa'):
                    c1, c2 = st.columns(2)
                    with c1:
                        novo_nome_t = st.text_input('Nome da atividade',
                                                    value=tarefa_sel.get('nome', ''))
                        nova_nota   = st.number_input('Nota', 0.0, 10.0,
                                                      value=float(tarefa_sel.get('nota', 0) or 0),
                                                      step=0.1)
                    with c2:
                        nova_prio = st.selectbox('Prioridade',
                                                 ['🔴 Alta', '🟡 Média', '🟢 Baixa'],
                                                 index=['🔴 Alta', '🟡 Média', '🟢 Baixa'].index(
                                                     tarefa_sel.get('prioridade', '🟡 Média')
                                                 ) if tarefa_sel.get('prioridade') in ['🔴 Alta', '🟡 Média', '🟢 Baixa'] else 1)
                        from datetime import datetime as _dt_edit
                        prazo_atual = tarefa_sel.get('prazo')
                        try:
                            prazo_val = _dt_edit.fromisoformat(str(prazo_atual)[:10]).date() if prazo_atual and str(prazo_atual) not in ('None','') else None
                        except Exception:
                            prazo_val = None
                        novo_prazo = st.date_input('Prazo', value=prazo_val)

                    if st.form_submit_button('💾 Salvar alterações', type='primary'):
                        res = api_patch('tarefas', id_edit, {
                            'nome':       novo_nome_t,
                            'nota':       float(nova_nota),
                            'prioridade': nova_prio,
                            'prazo':      novo_prazo.isoformat() if novo_prazo else None,
                        })
                        if res and res.status_code in (200, 201, 204):
                            st.success('✅ Atividade atualizada!')
                            _invalidar_cache()
                            st.rerun()
                        else:
                            st.error('Erro ao atualizar atividade.')

    with tab_remover:
        if not todas_tarefas:
            st.info('Nenhuma atividade cadastrada ainda.')
        else:
            id_del_t = st.selectbox(
                'Selecione a atividade para remover',
                options=[t['id'] for t in todas_tarefas],
                format_func=lambda x: next(
                    (t['nome'] for t in todas_tarefas if t['id'] == x), str(x)
                ),
                key='sel_remover_tarefa',
            )
            if st.button('🗑️ Remover Atividade', type='primary'):
                res = api_delete('tarefas', id_del_t)
                if res and res.status_code in (200, 204):
                    st.success('Atividade removida.')
                    _invalidar_cache()
                    st.rerun()
                else:
                    st.error('Erro ao remover atividade.')


# ------------------------------------------------
# 7. GERADOR DE RELATÓRIO PDF
# ------------------------------------------------

def gerar_pdf(df_notas: pd.DataFrame, df_metricas: dict, eventos: list = None, metas: list = None, anotacoes: list = None) -> bytes:
    '''
    Gera o relatório PDF completo do aluno usando ReportLab.
    Seções do PDF:
      1. Cabeçalho — logo EduTrack, data e nome do aluno
      2. Visão Geral — cards de métricas + barra de progresso
      3. Atividades — tabela com notas coloridas por faixa
      4. Agenda — eventos do calendário ordenados por data
      5. Metas — tabela de metas vs médias reais
      6. Anotações — bloco por disciplina
      7. Rodapé — identificação do relatório
    Retorna bytes prontos para download via st.download_button.
    '''
    """Gera relatório PDF com visual moderno e conteúdo rico."""
    from reportlab.platypus import HRFlowable, KeepTogether
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=1.8*cm, leftMargin=1.8*cm,
        topMargin=1.5*cm, bottomMargin=1.8*cm,
    )

    # ── Paleta de cores ──────────────────────────────────
    AMARELO   = colors.HexColor('#f5c842')
    AMARELO_D = colors.HexColor('#92400e')
    ESCURO    = colors.HexColor('#0f172a')
    CINZA_T   = colors.HexColor('#475569')
    CINZA_L   = colors.HexColor('#f1f5f9')
    BORDA     = colors.HexColor('#e2e8f0')
    VERDE     = colors.HexColor('#166534')
    VERDE_L   = colors.HexColor('#dcfce7')
    AMARELO_L = colors.HexColor('#fef9e7')
    VERMELHO  = colors.HexColor('#991b1b')
    VERMELHO_L= colors.HexColor('#fee2e2')
    BRANCO    = colors.white

    # ── Estilos ──────────────────────────────────────────
    s_nome    = ParagraphStyle('nome',    fontSize=22, fontName='Helvetica-Bold',
                               textColor=ESCURO, spaceAfter=8)
    s_sub_t   = ParagraphStyle('sub_t',   fontSize=11, fontName='Helvetica',
                               textColor=CINZA_T, spaceAfter=14)
    s_secao   = ParagraphStyle('secao',   fontSize=12, fontName='Helvetica-Bold',
                               textColor=ESCURO, spaceBefore=18, spaceAfter=8)
    s_normal  = ParagraphStyle('normal',  fontSize=9,  fontName='Helvetica',
                               textColor=CINZA_T, leading=14)
    s_rodape  = ParagraphStyle('rodape',  fontSize=8,  fontName='Helvetica',
                               textColor=colors.HexColor('#94a3b8'),
                               alignment=TA_CENTER)
    s_badge_v = ParagraphStyle('badge_v', fontSize=8, fontName='Helvetica-Bold',
                               textColor=VERDE,    alignment=TA_CENTER)
    s_badge_a = ParagraphStyle('badge_a', fontSize=8, fontName='Helvetica-Bold',
                               textColor=AMARELO_D, alignment=TA_CENTER)
    s_badge_r = ParagraphStyle('badge_r', fontSize=8, fontName='Helvetica-Bold',
                               textColor=VERMELHO, alignment=TA_CENTER)

    nome_usuario = df_metricas.get('nome_usuario', 'Estudante')
    story = []

    # ── CABEÇALHO ────────────────────────────────────────
    cabecalho_data = [[
        Paragraph(f'<b>EduTrack AI</b>', ParagraphStyle('logo', fontSize=14,
                  fontName='Helvetica-Bold', textColor=AMARELO_D)),
        Paragraph(f"Gerado em {date.today().strftime('%d/%m/%Y')}",
                  ParagraphStyle('data_r', fontSize=9, fontName='Helvetica',
                                 textColor=CINZA_T, alignment=TA_RIGHT)),
    ]]
    tabela_cab = Table(cabecalho_data, colWidths=[9*cm, 8*cm])
    tabela_cab.setStyle(TableStyle([
        ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING',(0,0), (-1,-1), 8),
        ('LINEBELOW',    (0,0), (-1,0), 1.5, AMARELO),
    ]))
    story.append(tabela_cab)
    story.append(Spacer(1, 0.4*cm))

    # Nome e subtítulo — separados para não sobrepor
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph('Relatório de Desempenho', s_nome))
    story.append(Spacer(1, 0.15*cm))
    story.append(Paragraph(
        f'{nome_usuario}  ·  {date.today().strftime("%B de %Y").capitalize()}',
        s_sub_t,
    ))

    # ── CARDS DE MÉTRICAS (linha de 5 cards) ─────────────
    story.append(Paragraph('Visão Geral', s_secao))

    def card_metrica(label, valor, destaque=False):
        bg = AMARELO if destaque else CINZA_L
        tc = AMARELO_D if destaque else ESCURO
        return [
            Paragraph(f'<b>{valor}</b>',
                      ParagraphStyle('cv', fontSize=18, fontName='Helvetica-Bold',
                                     textColor=tc, alignment=TA_CENTER)),
            Paragraph(label,
                      ParagraphStyle('cl', fontSize=7, fontName='Helvetica',
                                     textColor=CINZA_T, alignment=TA_CENTER)),
        ]

    media = df_metricas.get('media_geral', 0)
    conc  = df_metricas.get('concluidas', 0)
    total = df_metricas.get('total_tarefas', 0)
    pct   = int(conc / total * 100) if total > 0 else 0

    cards_data = [[
        card_metrica('DISCIPLINAS',  str(df_metricas.get('total_disciplinas', 0))),
        card_metrica('ATIVIDADES',   str(total)),
        card_metrica('CONCLUÍDAS',   str(conc)),
        card_metrica('PENDENTES',    str(df_metricas.get('pendentes', 0))),
        card_metrica('MÉDIA GERAL',  f"{media:.1f}", destaque=True),
    ]]
    # Cada card é uma sub-tabela
    cards_cells = []
    for card in cards_data[0]:
        sub = Table([[card[0]], [card[1]]], colWidths=[3.2*cm])
        sub.setStyle(TableStyle([
            ('BACKGROUND',   (0,0), (-1,-1), CINZA_L),
            ('TOPPADDING',   (0,0), (-1,-1), 8),
            ('BOTTOMPADDING',(0,0), (-1,-1), 8),
            ('ROUNDEDCORNERS', [4]),
        ]))
        cards_cells.append(sub)

    tabela_cards = Table([cards_cells], colWidths=[3.2*cm]*5,
                         hAlign='LEFT', spaceBefore=4)
    tabela_cards.setStyle(TableStyle([
        ('ALIGN',        (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
        ('COLPADDING',   (0,0), (-1,-1), 4),
    ]))
    story.append(tabela_cards)

    # ── BARRA DE PROGRESSO GERAL ─────────────────────────
    story.append(Spacer(1, 0.4*cm))
    prog_data = [[
        Paragraph(f'Progresso geral: <b>{pct}%</b> concluído',
                  ParagraphStyle('prog', fontSize=9, fontName='Helvetica',
                                 textColor=CINZA_T)),
    ]]
    prog_table = Table(prog_data, colWidths=[17*cm])
    prog_table.setStyle(TableStyle([
        ('BACKGROUND',   (0,0), (-1,-1), CINZA_L),
        ('LEFTPADDING',  (0,0), (-1,-1), 8),
        ('TOPPADDING',   (0,0), (-1,-1), 6),
        ('BOTTOMPADDING',(0,0), (-1,-1), 6),
        ('LINEBELOW',    (0,0), (-1,-1), 4, AMARELO if pct > 0 else BORDA),
    ]))
    story.append(prog_table)

    # ── TABELA DE ATIVIDADES ─────────────────────────────
    if not df_notas.empty:
        story.append(Paragraph('Atividades', s_secao))

        header_row = [
            Paragraph('<b>Atividade</b>',   ParagraphStyle('th', fontSize=8, fontName='Helvetica-Bold', textColor=BRANCO)),
            Paragraph('<b>Disciplina</b>',  ParagraphStyle('th', fontSize=8, fontName='Helvetica-Bold', textColor=BRANCO)),
            Paragraph('<b>Nota</b>',        ParagraphStyle('th', fontSize=8, fontName='Helvetica-Bold', textColor=BRANCO, alignment=TA_CENTER)),
            Paragraph('<b>Prazo</b>',       ParagraphStyle('th', fontSize=8, fontName='Helvetica-Bold', textColor=BRANCO, alignment=TA_CENTER)),
            Paragraph('<b>Status</b>',      ParagraphStyle('th', fontSize=8, fontName='Helvetica-Bold', textColor=BRANCO, alignment=TA_CENTER)),
        ]
        dados = [header_row]
        estilos = [
            ('BACKGROUND',    (0,0), (-1,0),  ESCURO),
            ('FONTSIZE',      (0,0), (-1,-1), 8),
            ('GRID',          (0,0), (-1,-1), 0.3, BORDA),
            ('TOPPADDING',    (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING',   (0,0), (-1,-1), 6),
            ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ]

        for i, (_, row) in enumerate(df_notas.iterrows(), start=1):
            nota_val  = float(row.get('nota', 0) or 0)
            concluida = bool(row.get('concluida', False))
            prazo_str = str(row.get('prazo', '') or '')

            # Cor da linha por nota
            if nota_val >= 7:
                bg_linha = VERDE_L
                s_nota   = s_badge_v
            elif nota_val >= 5:
                bg_linha = AMARELO_L
                s_nota   = s_badge_a
            else:
                bg_linha = VERMELHO_L
                s_nota   = s_badge_r

            # Prazo formatado
            try:
                from datetime import datetime as _dt_p
                prazo_dt  = _dt_p.fromisoformat(prazo_str[:10])
                prazo_fmt = prazo_dt.strftime('%d/%m/%Y')
            except Exception:
                prazo_fmt = '—'

            status_txt = 'Concluída' if concluida else 'Pendente'
            s_status   = s_badge_v   if concluida else s_badge_a

            dados.append([
                Paragraph(str(row.get('nome', ''))[:45],
                          ParagraphStyle('td', fontSize=8, fontName='Helvetica', textColor=ESCURO)),
                Paragraph(str(row.get('disciplina', ''))[:30],
                          ParagraphStyle('td', fontSize=8, fontName='Helvetica', textColor=CINZA_T)),
                Paragraph(f'<b>{nota_val:.1f}</b>', s_nota),
                Paragraph(prazo_fmt,
                          ParagraphStyle('td_c', fontSize=8, fontName='Helvetica',
                                         textColor=CINZA_T, alignment=TA_CENTER)),
                Paragraph(status_txt, s_status),
            ])
            estilos.append(('BACKGROUND', (0,i), (-1,i), bg_linha))
            # Linha alternada mais clara
            if i % 2 == 0:
                estilos.append(('BACKGROUND', (0,i), (-1,i),
                                colors.HexColor('#f8fafc') if bg_linha == VERDE_L else bg_linha))

        tabela_notas = Table(dados, colWidths=[6.5*cm, 4.5*cm, 1.8*cm, 2.5*cm, 2.1*cm])
        tabela_notas.setStyle(TableStyle(estilos))
        story.append(tabela_notas)

    # ── AGENDA ───────────────────────────────────────────
    if eventos:
        from datetime import datetime as _dt_pdf
        story.append(Paragraph('Agenda de Compromissos', s_secao))

        CORES_TIPO = {
            '📝 Prova':        colors.HexColor('#fee2e2'),
            '📌 Entrega':      colors.HexColor('#fef9e7'),
            '🎯 Apresentação': colors.HexColor('#ede9fe'),
            '📖 Outro':        CINZA_L,
        }
        TEXTO_TIPO = {
            '📝 Prova':        VERMELHO,
            '📌 Entrega':      AMARELO_D,
            '🎯 Apresentação': colors.HexColor('#4338ca'),
            '📖 Outro':        CINZA_T,
        }

        eventos_ord = []
        for ev in eventos:
            try:
                d = _dt_pdf.fromisoformat(
                    str(ev.get('data_evento', ev.get('data', '')))[:10]
                )
                eventos_ord.append((d, ev))
            except Exception:
                pass
        eventos_ord.sort(key=lambda x: x[0])

        if eventos_ord:
            agenda_header = [
                Paragraph('<b>Data</b>',   ParagraphStyle('ath', fontSize=8, fontName='Helvetica-Bold', textColor=BRANCO)),
                Paragraph('<b>Tipo</b>',   ParagraphStyle('ath', fontSize=8, fontName='Helvetica-Bold', textColor=BRANCO)),
                Paragraph('<b>Evento</b>', ParagraphStyle('ath', fontSize=8, fontName='Helvetica-Bold', textColor=BRANCO)),
            ]
            agenda_dados  = [agenda_header]
            agenda_estilos = [
                ('BACKGROUND',    (0,0), (-1,0),  ESCURO),
                ('FONTSIZE',      (0,0), (-1,-1), 8),
                ('GRID',          (0,0), (-1,-1), 0.3, BORDA),
                ('TOPPADDING',    (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
                ('LEFTPADDING',   (0,0), (-1,-1), 6),
                ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
            ]
            for i, (d, ev) in enumerate(eventos_ord, start=1):
                tipo    = ev.get('tipo', '📖 Outro')
                nome_ev = str(ev.get('nome', ''))[:50]
                bg_ev   = CORES_TIPO.get(tipo, CINZA_L)
                tc_ev   = TEXTO_TIPO.get(tipo, CINZA_T)
                agenda_dados.append([
                    Paragraph(d.strftime('%d/%m/%Y'),
                              ParagraphStyle('adate', fontSize=8, fontName='Helvetica',
                                             textColor=CINZA_T, alignment=TA_CENTER)),
                    Paragraph(tipo.split(' ', 1)[-1],
                              ParagraphStyle('atipo', fontSize=8, fontName='Helvetica-Bold',
                                             textColor=tc_ev)),
                    Paragraph(nome_ev,
                              ParagraphStyle('anome', fontSize=8, fontName='Helvetica',
                                             textColor=ESCURO)),
                ])
                agenda_estilos.append(('BACKGROUND', (0,i), (-1,i), bg_ev))

            tabela_agenda = Table(agenda_dados, colWidths=[2.5*cm, 3.5*cm, 11.4*cm])
            tabela_agenda.setStyle(TableStyle(agenda_estilos))
            story.append(tabela_agenda)

    # ── METAS ────────────────────────────────────────────
    if metas and not df_notas.empty:
        story.append(Paragraph('Metas de Desempenho', s_secao))

        nomes_disc_pdf = {}
        if not df_notas.empty and 'disciplina' in df_notas.columns and 'disc_id' in df_notas.columns:
            nomes_disc_pdf = dict(zip(df_notas['disc_id'], df_notas['disciplina']))

        meta_header = [
            Paragraph('<b>Disciplina</b>', ParagraphStyle('mth', fontSize=8, fontName='Helvetica-Bold', textColor=BRANCO)),
            Paragraph('<b>Meta</b>',       ParagraphStyle('mth', fontSize=8, fontName='Helvetica-Bold', textColor=BRANCO, alignment=TA_CENTER)),
            Paragraph('<b>Média atual</b>',ParagraphStyle('mth', fontSize=8, fontName='Helvetica-Bold', textColor=BRANCO, alignment=TA_CENTER)),
            Paragraph('<b>Situação</b>',   ParagraphStyle('mth', fontSize=8, fontName='Helvetica-Bold', textColor=BRANCO, alignment=TA_CENTER)),
        ]
        meta_dados  = [meta_header]
        meta_estilos = [
            ('BACKGROUND',    (0,0), (-1,0),  ESCURO),
            ('FONTSIZE',      (0,0), (-1,-1), 8),
            ('GRID',          (0,0), (-1,-1), 0.3, BORDA),
            ('TOPPADDING',    (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING',   (0,0), (-1,-1), 6),
            ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ]

        # Calcula média por disciplina
        medias_disc = {}
        if not df_notas.empty and 'nota' in df_notas.columns and 'disc_id' in df_notas.columns:
            medias_disc = df_notas.groupby('disc_id')['nota'].mean().to_dict()

        for i, m in enumerate(metas, start=1):
            disc_id    = m.get('disc_id')
            disc_nome  = nomes_disc_pdf.get(disc_id, f'Disciplina {disc_id}')
            valor_meta = float(m.get('valor_meta', 7.0))
            media_disc = float(medias_disc.get(disc_id, 0.0))
            atingida   = media_disc >= valor_meta

            situacao_txt = 'Atingida ✓' if atingida else f'Faltam {valor_meta - media_disc:.1f}'
            bg_meta  = VERDE_L   if atingida else AMARELO_L
            tc_meta  = VERDE     if atingida else AMARELO_D

            meta_dados.append([
                Paragraph(disc_nome, ParagraphStyle('mtd', fontSize=8, fontName='Helvetica', textColor=ESCURO)),
                Paragraph(f'{valor_meta:.1f}', ParagraphStyle('mtd_c', fontSize=8, fontName='Helvetica-Bold', textColor=CINZA_T, alignment=TA_CENTER)),
                Paragraph(f'{media_disc:.1f}', ParagraphStyle('mtd_c', fontSize=8, fontName='Helvetica-Bold', textColor=tc_meta, alignment=TA_CENTER)),
                Paragraph(situacao_txt, ParagraphStyle('mtd_s', fontSize=8, fontName='Helvetica-Bold', textColor=tc_meta, alignment=TA_CENTER)),
            ])
            meta_estilos.append(('BACKGROUND', (0,i), (-1,i), bg_meta))

        tabela_metas = Table(meta_dados, colWidths=[8*cm, 2.5*cm, 3*cm, 4*cm])
        tabela_metas.setStyle(TableStyle(meta_estilos))
        story.append(tabela_metas)

    # ── ANOTAÇÕES ────────────────────────────────────────
    if anotacoes:
        anot_com_texto = [a for a in anotacoes if str(a.get('texto', '')).strip()]
        if anot_com_texto:
            story.append(Paragraph('Anotações por Disciplina', s_secao))

            nomes_disc_pdf = {}
            if not df_notas.empty and 'disciplina' in df_notas.columns and 'disc_id' in df_notas.columns:
                nomes_disc_pdf = dict(zip(df_notas['disc_id'], df_notas['disciplina']))

            for anot in anot_com_texto:
                disc_id   = anot.get('disc_id')
                disc_nome = nomes_disc_pdf.get(disc_id, f'Disciplina {disc_id}')
                texto     = str(anot.get('texto', '')).strip()

                anot_bloco = Table([
                    [Paragraph(f'<b>{disc_nome}</b>',
                               ParagraphStyle('an_h', fontSize=9, fontName='Helvetica-Bold',
                                              textColor=AMARELO_D))],
                    [Paragraph(texto[:800],
                               ParagraphStyle('an_t', fontSize=8, fontName='Helvetica',
                                              textColor=CINZA_T, leading=13))],
                ], colWidths=[17.4*cm])
                anot_bloco.setStyle(TableStyle([
                    ('BACKGROUND',    (0,0), (-1,0), AMARELO_L),
                    ('BACKGROUND',    (0,1), (-1,-1), CINZA_L),
                    ('LEFTPADDING',   (0,0), (-1,-1), 10),
                    ('RIGHTPADDING',  (0,0), (-1,-1), 10),
                    ('TOPPADDING',    (0,0), (-1,-1), 6),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                    ('LINEBELOW',     (0,0), (-1,0), 0.5, AMARELO),
                    ('ROUNDEDCORNERS', [4]),
                ]))
                story.append(anot_bloco)
                story.append(Spacer(1, 0.2*cm))

    # ── RODAPÉ ───────────────────────────────────────────
    story.append(Spacer(1, 0.8*cm))
    story.append(HRFlowable(width='100%', thickness=0.5,
                             color=BORDA, spaceAfter=6))
    story.append(Paragraph(
        f'EduTrack AI  ·  Relatório gerado automaticamente em {date.today().strftime("%d/%m/%Y")}  ·  {nome_usuario}',
        s_rodape,
    ))

    doc.build(story)
    buffer.seek(0)
    return buffer.read()




# ------------------------------------------------
# 8. DASHBOARD
# ------------------------------------------------

def modulo_dashboard():
    '''
    Dashboard principal — primeira tela após o login.
    Exibe métricas, gráficos de desempenho e alertas de prazo.
    Todos os dados são buscados do Xano e processados com pandas.
    '''
    from datetime import datetime
    hora = datetime.now().hour
    saudacao = 'Bom dia' if hora < 12 else 'Boa tarde' if hora < 18 else 'Boa noite'
    nome_usuario = st.session_state.get('user_name', 'Estudante')
    hoje = datetime.now().strftime('%d/%m/%Y')
    st.header(f'{saudacao}, {nome_usuario} 👋')
    st.caption(f'Hoje é {hoje}')

    tarefas = api_get('tarefas', st.session_state.get('auth_token', ''))
    discs   = api_get('disciplinas', st.session_state.get('auth_token', ''))

    if not tarefas or not discs:
        st.markdown(f"""
        <div style="text-align:center;padding:60px 20px">
          <div style="font-size:52px;margin-bottom:16px">🚀</div>
          <div style="font-size:22px;font-weight:500;color:#f1f5f9;margin-bottom:10px">
            Vamos começar!
          </div>
          <div style="font-size:15px;color:#94a3b8;margin-bottom:32px;line-height:1.7">
            Cadastre um professor para liberar todas as<br>funcionalidades do EduTrack.
          </div>
        </div>
        """, unsafe_allow_html=True)
        col_a, col_b, col_c = st.columns([2, 1, 2])
        with col_b:
            if st.button('+ Cadastrar professor', use_container_width=True):
                st.session_state['_nav_target'] = 'Professores'
                st.rerun()
        return

    # FEATURE: alertas de vencimento no topo do dashboard
    alertas_vencimento(tarefas, discs)

    df_t = pd.DataFrame(tarefas)


    df_d = pd.DataFrame(discs)

    if 'concluida' not in df_t.columns:
        df_t['concluida'] = False

    nomes_disc           = {d['id']: d['nome'] for d in discs}
    df_t['disciplina']   = df_t['disc_id'].map(nomes_disc).fillna('—')

    total_disc    = len(df_d)
    total_tarefas = len(df_t)
    concluidas    = int(df_t['concluida'].sum())
    pendentes     = total_tarefas - concluidas
    media_geral   = float(df_t['nota'].mean()) if 'nota' in df_t.columns else 0.0

    # Cards de métricas
    melhor_nota = float(df_t['nota'].max()) if 'nota' in df_t.columns and total_tarefas > 0 else 0.0
    pior_nota   = float(df_t['nota'].min()) if 'nota' in df_t.columns and total_tarefas > 0 else 0.0

    st.subheader('Visão Geral')
    col1, col2, col3, col4, col5, col6, col7 = st.columns(7)
    col1.metric('Disciplinas',  total_disc)
    col2.metric('Atividades',   total_tarefas)
    col3.metric('Concluídas',   concluidas)
    col4.metric('Pendentes',    pendentes)
    col5.metric('Média Geral',  f'{media_geral:.1f}')
    col6.metric('Melhor nota',  f'{melhor_nota:.1f}',
                delta=f'+{melhor_nota - media_geral:.1f}' if melhor_nota > media_geral else None)
    col7.metric('Pior nota',    f'{pior_nota:.1f}',
                delta=f'{pior_nota - media_geral:.1f}' if pior_nota < media_geral else None,
                delta_color='normal')

    # Contador de atividades por disciplina
    st.markdown('---')
    st.subheader('Atividades por disciplina')
    nomes_disc_dash = {d['id']: d['nome'] for d in discs}
    df_t['disciplina'] = df_t['disc_id'].map(nomes_disc_dash).fillna('—')
    cols_disc = st.columns(min(len(discs), 4))
    for i, disc in enumerate(discs):
        disc_id   = disc['id']
        disc_nome = disc['nome']
        df_disc   = df_t[df_t['disc_id'] == disc_id]
        total_d   = len(df_disc)
        conc_d    = int(df_disc['concluida'].sum()) if 'concluida' in df_disc.columns else 0
        media_d   = float(df_disc['nota'].mean()) if 'nota' in df_disc.columns and total_d > 0 else 0.0
        with cols_disc[i % min(len(discs), 4)]:
            st.metric(
                label=disc_nome[:25] + ('…' if len(disc_nome) > 25 else ''),
                value=f'{total_d} atividade(s)',
                delta=f'Média {media_d:.1f} · {conc_d} concluída(s)',
                delta_color='normal' if media_d >= 7 else 'inverse',
            )

    st.markdown('---')

    import plotly.graph_objects as go

    df_plot = df_t.merge(df_d, left_on='disc_id', right_on='id', suffixes=('_t', '_d'))

    # ── Paleta e configurações base ─────────────────────────
    CORES  = ['#f5c842', '#f97316', '#06b6d4', '#8b5cf6', '#ec4899', '#10b981']
    FONTE  = 'Inter, sans-serif'

    # Tema fixo: dark
    TEMPLATE     = 'plotly_dark'
    COR_TEXTO    = 'rgba(255,255,255,0.75)'
    COR_TITULO   = '#ffffff'
    COR_GRID     = 'rgba(255,255,255,0.08)'
    COR_HOVER_BG = '#1e2130'
    COR_HOVER_TX = '#ffffff'
    FUNDO        = 'rgba(0,0,0,0)'

    def base_layout(titulo: str) -> dict:
        return dict(
            template=TEMPLATE,
            title=dict(
                text=f'<b>{titulo}</b>',
                font=dict(family=FONTE, size=14, color=COR_TITULO),
                x=0, xanchor='left', pad=dict(l=4, b=8),
            ),
            plot_bgcolor=FUNDO,
            paper_bgcolor=FUNDO,
            font=dict(family=FONTE, size=11, color=COR_TEXTO),
            margin=dict(l=4, r=4, t=48, b=4),
            hoverlabel=dict(
                bgcolor=COR_HOVER_BG,
                bordercolor='rgba(255,255,255,0.1)',
                font=dict(family=FONTE, size=12, color=COR_HOVER_TX),
            ),
        )

    def eixo(titulo='', angulo=0, grid=True) -> dict:
        return dict(
            title=dict(text=titulo, font=dict(size=11, color=COR_TEXTO)),
            tickfont=dict(size=11, color=COR_TEXTO, family=FONTE),
            tickangle=angulo,
            gridcolor=COR_GRID if grid else FUNDO,
            gridwidth=1,
            linecolor='rgba(255,255,255,0.1)',
            zeroline=False,
            showgrid=grid,
        )

    # ═══════════════════════════════════════════════════════
    # LINHA 1 — Notas por atividade | Média por disciplina
    # ═══════════════════════════════════════════════════════
    col_g1, col_g2 = st.columns(2)

    with col_g1:
        fig_bar = px.bar(
            df_plot,
            x='nome_t', y='nota', color='nome_d',
            barmode='group',
            labels={'nome_t': '', 'nota': 'Nota', 'nome_d': ''},
            color_discrete_sequence=CORES,
            custom_data=['nome_d'],
        )
        fig_bar.update_traces(
            hovertemplate='<b>%{x}</b><br>%{customdata[0]}<br>Nota: <b>%{y:.1f}</b><extra></extra>',
            marker_line_width=0,
            texttemplate='%{y:.1f}',
            textposition='outside',
            textfont=dict(size=10, color=COR_TEXTO, family=FONTE),
        )
        fig_bar.update_layout(
            **base_layout('Notas por atividade'),
            xaxis=dict(**eixo(angulo=-30, grid=False), automargin=True),
            yaxis=dict(**eixo(grid=True), range=[0, 12]),
            bargap=0.2, bargroupgap=0.05,
            legend=dict(
                bgcolor=FUNDO, font=dict(size=11, color=COR_TEXTO),
                orientation='h', yanchor='bottom', y=1.01,
                xanchor='right', x=1, title_text='',
            ),
            showlegend=True,
        )
        st.plotly_chart(fig_bar, use_container_width=True)

    with col_g2:
        df_media = (
            df_plot.groupby('nome_d', as_index=False)['nota']
            .mean()
            .rename(columns={'nome_d': 'Disciplina', 'nota': 'Média'})
            .sort_values('Média', ascending=True)
        )
        cores_media = [
            '#f87171' if m < 5 else '#fbbf24' if m < 7 else '#34d399'
            for m in df_media['Média']
        ]
        fig_horiz = go.Figure(go.Bar(
            x=df_media['Média'],
            y=df_media['Disciplina'],
            orientation='h',
            marker=dict(
                color=cores_media,
                line=dict(width=0),
                opacity=0.9,
            ),
            text=[f'{m:.1f}' for m in df_media['Média']],
            textposition='outside',
            textfont=dict(size=12, color=COR_TEXTO, family=FONTE),
            hovertemplate='<b>%{y}</b><br>Média: <b>%{x:.1f}</b><extra></extra>',
            width=0.5,
        ))
        fig_horiz.update_layout(
            **base_layout('Média por disciplina'),
            xaxis=dict(**eixo(grid=True), range=[0, 12]),
            yaxis=dict(**eixo(grid=False), automargin=True),
            showlegend=False,
        )
        st.plotly_chart(fig_horiz, use_container_width=True)

    # ═══════════════════════════════════════════════════════
    # LINHA 2 — Evolução temporal (largura total)
    # ═══════════════════════════════════════════════════════
    modulo_evolucao(df_t, df_d, COR_TEXTO, COR_TITULO, COR_GRID, TEMPLATE, FONTE, COR_HOVER_BG, COR_HOVER_TX)

    # ═══════════════════════════════════════════════════════
    # LINHA 3 — Radar 360° por disciplina
    # ═══════════════════════════════════════════════════════
    if len(discs) >= 2:  # Radar só faz sentido com 2+ disciplinas
        col_radar, col_info = st.columns([2, 1])

        with col_radar:
            # Monta os dados do radar para cada disciplina
            categorias = ['Média (0-10)', '% Concluídas', 'Atividades lançadas']

            # Normaliza "Atividades lançadas" para escala 0-10
            # para ficar comparável com os outros eixos
            max_ativs = max(
                len(df_t[df_t['disc_id'] == d['id']]) for d in discs
            ) or 1

            fig_radar = go.Figure()

            for i, disc in enumerate(discs):
                disc_id   = disc['id']
                disc_nome = disc['nome']
                df_disc   = df_t[df_t['disc_id'] == disc_id]

                if df_disc.empty:
                    continue

                # Calcula as 3 dimensões
                media_disc = float(df_disc['nota'].mean()) if 'nota' in df_disc.columns else 0.0

                if 'concluida' in df_disc.columns:
                    df_disc = df_disc.copy()
                    df_disc['concluida'] = df_disc['concluida'].fillna(False).astype(bool)
                    pct_conc = float(df_disc['concluida'].mean() * 10)  # 0-10
                else:
                    pct_conc = 0.0

                qtd_norm = float(len(df_disc) / max_ativs * 10)  # 0-10 normalizado

                valores = [media_disc, pct_conc, qtd_norm]
                # Fecha o polígono repetindo o primeiro valor
                valores_fechados   = valores + [valores[0]]
                categorias_fechadas = categorias + [categorias[0]]

                cor = CORES[i % len(CORES)]

                # Converte hex para rgba para o fillcolor
                try:
                    r_c = int(cor[1:3], 16)
                    g_c = int(cor[3:5], 16)
                    b_c = int(cor[5:7], 16)
                    fill_cor = f'rgba({r_c},{g_c},{b_c},0.18)'
                except Exception:
                    fill_cor = 'rgba(245,200,66,0.18)'

                # Abreviação do nome para o label inline
                palavras   = disc_nome.split()
                abrev      = ' '.join(p[:4] for p in palavras[:3])
                label_nome = abrev if len(disc_nome) > 14 else disc_nome

                fig_radar.add_trace(go.Scatterpolar(
                    r=valores_fechados,
                    theta=categorias_fechadas,
                    fill='toself',
                    fillcolor=fill_cor,
                    line=dict(color=cor, width=2.5),
                    name=disc_nome,          # legenda completa no hover
                    hovertemplate=(
                        f'<b>{disc_nome}</b><br>'
                        '%{theta}: <b>%{r:.1f}</b><extra></extra>'
                    ),
                ))

                # Label do nome da disciplina — posicionado no vértice de maior valor
                idx_max   = valores.index(max(valores))
                # Ângulos dos 3 vértices do radar (0°, 120°, 240°)
                angulos   = [90, 210, 330]  # posições angulares dos eixos
                import math
                ang_rad   = math.radians(angulos[idx_max])
                raio_lbl  = min(max(valores) + 1.5, 10.5)
                # Converte polar → cartesiano para posicionar a anotação
                x_lbl = 0.5 + (raio_lbl / 14) * math.cos(ang_rad)
                y_lbl = 0.5 + (raio_lbl / 14) * math.sin(ang_rad)

                fig_radar.add_annotation(
                    text=f'<b>{label_nome}</b>',
                    x=x_lbl, y=y_lbl,
                    xref='paper', yref='paper',
                    showarrow=False,
                    font=dict(size=11, color=cor, family=FONTE),
                    bgcolor='rgba(0,0,0,0)',
                    borderwidth=0,
                    xanchor='center',
                )

            fig_radar.update_layout(
                template=TEMPLATE,
                title=dict(
                    text='<b>Visão 360° por disciplina</b>',
                    font=dict(family=FONTE, size=14, color=COR_TITULO),
                    x=0, xanchor='left', pad=dict(l=4, b=8),
                ),
                polar=dict(
                    bgcolor=FUNDO,
                    radialaxis=dict(
                        visible=True,
                        range=[0, 10],
                        tickfont=dict(size=9, color=COR_TEXTO),
                        gridcolor=COR_GRID,
                        linecolor=COR_GRID,
                        tickvals=[0, 2.5, 5, 7.5, 10],
                    ),
                    angularaxis=dict(
                        tickfont=dict(size=11, color=COR_TEXTO, family=FONTE),
                        gridcolor=COR_GRID,
                        linecolor=COR_GRID,
                    ),
                ),
                plot_bgcolor=FUNDO,
                paper_bgcolor=FUNDO,
                font=dict(family=FONTE, size=11, color=COR_TEXTO),
                margin=dict(l=4, r=4, t=48, b=4),
                legend=dict(
                    bgcolor=FUNDO,
                    font=dict(size=12, color=COR_TEXTO, family=FONTE),
                    orientation='h',
                    yanchor='top',   y=-0.08,
                    xanchor='center', x=0.5,
                    title_text='',
                ),
                hoverlabel=dict(
                    bgcolor=COR_HOVER_BG,
                    font=dict(color=COR_HOVER_TX, size=12, family=FONTE),
                ),
            )
            st.plotly_chart(fig_radar, use_container_width=True)

        with col_info:
            # Legenda explicativa dos eixos do radar
            st.markdown('**Como ler o radar**')
            st.markdown('---')
            infos = [
                ('📊', 'Média (0-10)',
                 'Média aritmética de todas as notas da disciplina.'),
                ('✅', '% Concluídas',
                 'Proporção de atividades marcadas como concluídas (0 = nenhuma, 10 = todas).'),
                ('📝', 'Atividades lançadas',
                 'Quantidade relativa de atividades — normalizada em relação à disciplina com mais atividades.'),
            ]
            for icone, titulo, desc in infos:
                st.markdown(f'**{icone} {titulo}**')
                st.caption(desc)
                st.markdown('')

            st.markdown('---')
            st.caption(
                '💡 Quanto maior a área coberta, '
                'melhor o desempenho geral na disciplina.'
            )

    # Botão PDF direto no dashboard
    st.markdown('---')
    st.subheader('📄 Relatório')
    st.caption('Baixe um PDF completo com suas notas e métricas.')

    metricas = {
        'total_disciplinas': total_disc,
        'total_tarefas':     total_tarefas,
        'concluidas':        concluidas,
        'pendentes':         pendentes,
        'media_geral':       media_geral,
        'nome_usuario':      st.session_state.get('user_name', 'Estudante'),
    }

    if st.button('📄 Gerar Relatório PDF', type='primary'):
        with st.spinner('Gerando PDF...'):
            eventos_pdf   = api_get('eventos', st.session_state.get('auth_token', ''))   or []
            metas_pdf     = api_get('metas', st.session_state.get('auth_token', ''))      or []
            anotacoes_pdf = api_get('anotacoes', st.session_state.get('auth_token', ''))  or []
            pdf_bytes     = gerar_pdf(df_t, metricas,
                                      eventos=eventos_pdf,
                                      metas=metas_pdf,
                                      anotacoes=anotacoes_pdf)
        st.download_button(
            label='⬇️ Baixar PDF',
            data=pdf_bytes,
            file_name=f"edutrack_relatorio_{date.today().isoformat()}.pdf",
            mime='application/pdf',
        )



# ------------------------------------------------
# 9. MÓDULO CALENDÁRIO
# ------------------------------------------------

def modulo_calendario():
    '''
    Calendário mensal com eventos visuais.
    Funcionalidades:
      - Grade HTML gerada dinamicamente com dias do mês atual
      - Pontinhos coloridos nos dias com eventos (por tipo)
      - Lista de próximos eventos com countdown de dias
      - CRUD completo: criar, editar e remover eventos
      - Tipos: Prova (vermelho), Entrega (amarelo),
        Apresentação (roxo), Outro (cinza)
    Os eventos são salvos no Xano filtrados por user_id.
    '''
    from datetime import datetime, timedelta
    import calendar as cal_mod

    MESES_PT = {
        1:'Janeiro', 2:'Fevereiro', 3:'Março', 4:'Abril',
        5:'Maio', 6:'Junho', 7:'Julho', 8:'Agosto',
        9:'Setembro', 10:'Outubro', 11:'Novembro', 12:'Dezembro'
    }
    DIAS_PT = ['Dom','Seg','Ter','Qua','Qui','Sex','Sáb']

    st.header('📅 Calendário')
    st.caption('Registre datas de provas, entregas e compromissos')

    # Navegação de meses
    hoje = datetime.today()
    if 'cal_ano' not in st.session_state:
        st.session_state.cal_ano = hoje.year
    if 'cal_mes' not in st.session_state:
        st.session_state.cal_mes = hoje.month

    # ── Formulário ─────────────────────────────────────
    with st.expander('➕  Novo evento'):
        with st.form('form_add_evento'):
            c1, c2, c3 = st.columns([3, 1, 2])
            with c1:
                nome_ev = st.text_input('Nome do evento', placeholder='Ex: Prova de Matemática')
            with c2:
                data_ev = st.date_input('Data', value=datetime.today())
            with c3:
                tipo_ev = st.selectbox('Tipo', ['📝 Prova', '📌 Entrega', '🎯 Apresentação', '📖 Outro'])
            if st.form_submit_button('Salvar Evento', type='primary'):
                if nome_ev:
                    res = api_post('eventos', {
                        'nome':        nome_ev,
                        'data_evento': data_ev.isoformat(),  # campo renomeado no Xano
                        'tipo':        tipo_ev,
                    })
                    if res and res.status_code in (200, 201):
                        st.success('Evento salvo!')
                        _invalidar_cache()
                        st.rerun()
                    else:
                        st.error('Erro ao salvar evento.')
                else:
                    st.warning('Digite o nome do evento.')

    eventos = api_get('eventos', st.session_state.get('auth_token', '')) or []

    hoje    = datetime.today()
    ano     = st.session_state.cal_ano
    mes     = st.session_state.cal_mes
    nome_mes = f"{MESES_PT[mes]} de {ano}"

    # Datas com eventos no mês atual
    datas_eventos = {}      # dia → lista de tipos
    for ev in eventos:
        try:
            d = datetime.fromisoformat(str(ev.get('data_evento', ev.get('data', '')))[:10])
            if d.year == ano and d.month == mes:
                datas_eventos.setdefault(d.day, []).append(ev.get('tipo',''))
        except Exception:
            pass

    # ── Layout principal: calendário | eventos ─────────
    col_cal, col_ev = st.columns([1.2, 1])

    with col_cal:
        # Cabeçalho do mês com navegação
        col_prev, col_titulo, col_next = st.columns([1, 4, 1])
        with col_prev:
            if st.button('◀', key='cal_prev', use_container_width=True):
                if st.session_state.cal_mes == 1:
                    st.session_state.cal_mes = 12
                    st.session_state.cal_ano -= 1
                else:
                    st.session_state.cal_mes -= 1
                st.rerun()
        with col_titulo:
            st.markdown(
                f'<div style="font-size:18px;font-weight:600;text-align:center;padding:8px 0">'
                f'{nome_mes}</div>',
                unsafe_allow_html=True,
            )
        with col_next:
            if st.button('▶', key='cal_next', use_container_width=True):
                if st.session_state.cal_mes == 12:
                    st.session_state.cal_mes = 1
                    st.session_state.cal_ano += 1
                else:
                    st.session_state.cal_mes += 1
                st.rerun()

        # Dias da semana
        header = ''.join(
            f'<div style="text-align:center;font-size:11px;font-weight:500;'
            f'color:#9ca3af;padding-bottom:8px;text-transform:uppercase;letter-spacing:.05em">'
            f'{d}</div>'
            for d in DIAS_PT
        )

        # Offset para domingo = 0
        primeiro_dia = datetime(ano, mes, 1)
        offset = (primeiro_dia.weekday() + 1) % 7
        total_dias = cal_mod.monthrange(ano, mes)[1]

        cells = ''
        for _ in range(offset):
            cells += '<div></div>'

        for dia in range(1, total_dias + 1):
            is_hoje = (dia == hoje.day)
            has_ev  = dia in datas_eventos

            if is_hoje:
                bg, cor_num, fw = '#f5c842', '#6b4f00', '700'
                borda = 'none'
            elif has_ev:
                bg, cor_num, fw = 'transparent', '#f5c842', '600'
                borda = '1.5px solid #f5c842'
            else:
                bg, cor_num, fw = 'transparent', 'inherit', '400'
                borda = 'none'

            # Pontinhos de tipo de evento embaixo do número
            dots = ''
            if has_ev:
                cores_tipo_dot = {
                    '📝 Prova':       '#f87171',
                    '📌 Entrega':     '#f5c842',
                    '🎯 Apresentação':'#8b5cf6',
                    '📖 Outro':       '#9ca3af',
                }
                for tipo in datas_eventos[dia][:3]:
                    cor_dot = cores_tipo_dot.get(tipo, '#9ca3af')
                    dots += f'<div style="width:4px;height:4px;border-radius:50%;background:{cor_dot};margin:0 1px"></div>'
                dots = f'<div style="display:flex;justify-content:center;margin-top:2px">{dots}</div>'

            cells += (
                f'<div style="text-align:center;padding:6px 2px 4px;border-radius:8px;'
                f'background:{bg};border:{borda};cursor:default;min-height:42px">'
                f'<div style="font-size:13px;font-weight:{fw};color:{cor_num};line-height:1">{dia}</div>'
                f'{dots}'
                f'</div>'
            )

        grade_html = (
            f'<div style="display:grid;grid-template-columns:repeat(7,1fr);gap:4px">'
            f'{header}{cells}</div>'
        )
        st.markdown(grade_html, unsafe_allow_html=True)

        # Legenda dos tipos
        st.markdown(
            '''<div style="display:flex;gap:14px;margin-top:16px;flex-wrap:wrap">
            <span style="font-size:11px;color:#9ca3af;display:flex;align-items:center;gap:4px">
              <span style="width:8px;height:8px;border-radius:50%;background:#f87171;display:inline-block"></span>Prova
            </span>
            <span style="font-size:11px;color:#9ca3af;display:flex;align-items:center;gap:4px">
              <span style="width:8px;height:8px;border-radius:50%;background:#f5c842;display:inline-block"></span>Entrega
            </span>
            <span style="font-size:11px;color:#9ca3af;display:flex;align-items:center;gap:4px">
              <span style="width:8px;height:8px;border-radius:50%;background:#8b5cf6;display:inline-block"></span>Apresentação
            </span>
            <span style="font-size:11px;color:#9ca3af;display:flex;align-items:center;gap:4px">
              <span style="width:8px;height:8px;border-radius:50%;background:#9ca3af;display:inline-block"></span>Outro
            </span>
            </div>''',
            unsafe_allow_html=True,
        )

    with col_ev:
        st.markdown(
            '<div style="font-size:18px;font-weight:600;margin-bottom:16px">Próximos eventos</div>',
            unsafe_allow_html=True,
        )

        if not eventos:
            st.caption('Nenhum evento cadastrado ainda.')
        else:
            df_ev = pd.DataFrame(eventos)
            try:
                df_ev['data_dt'] = pd.to_datetime(df_ev.get('data_evento', df_ev.get('data', '')).astype(str).str[:10], errors='coerce')
                df_ev = df_ev[df_ev['data_dt'] >= pd.Timestamp.today()].sort_values('data_dt')
            except Exception:
                df_ev = pd.DataFrame()

            if df_ev.empty:
                st.caption('Sem eventos futuros cadastrados.')
            else:
                cores_tipo_ev = {
                    '📝 Prova':        ('#f87171', 'rgba(248,113,113,0.12)'),
                    '📌 Entrega':      ('#f5c842', 'rgba(245,200,66,0.12)'),
                    '🎯 Apresentação': ('#8b5cf6', 'rgba(139,92,246,0.12)'),
                    '📖 Outro':        ('#9ca3af', 'rgba(156,163,175,0.10)'),
                }
                for _, row in df_ev.iterrows():
                    tipo     = row.get('tipo', '📖 Outro')
                    nome     = row.get('nome', '')
                    data_dt  = row.get('data_dt')
                    cor_ev, bg_ev = cores_tipo_ev.get(tipo, ('#9ca3af', 'rgba(156,163,175,0.10)'))

                    try:
                        data_fmt = data_dt.strftime('%d/%m')
                        dias_rest = (data_dt.date() - hoje.date()).days
                        if dias_rest == 0:
                            prazo_txt = 'Hoje'
                            prazo_cor = '#f87171'
                        elif dias_rest == 1:
                            prazo_txt = 'Amanhã'
                            prazo_cor = '#f5c842'
                        elif dias_rest <= 7:
                            prazo_txt = f'{dias_rest}d'
                            prazo_cor = '#f5c842'
                        else:
                            prazo_txt = f'{dias_rest}d'
                            prazo_cor = '#9ca3af'
                    except Exception:
                        data_fmt = '—'
                        prazo_txt = ''
                        prazo_cor = '#9ca3af'

                    st.markdown(
                        f'''<div style="display:flex;align-items:center;gap:12px;
                        padding:12px 14px;margin-bottom:8px;border-radius:10px;
                        background:{bg_ev};border-left:3px solid {cor_ev}">
                          <div style="flex:1;min-width:0">
                            <div style="font-size:13px;font-weight:500;
                                        white-space:nowrap;overflow:hidden;text-overflow:ellipsis">
                              {nome}
                            </div>
                            <div style="font-size:11px;color:#9ca3af;margin-top:2px">{tipo}</div>
                          </div>
                          <div style="text-align:right;flex-shrink:0">
                            <div style="font-size:13px;font-weight:500">{data_fmt}</div>
                            <div style="font-size:11px;color:{prazo_cor};font-weight:500">{prazo_txt}</div>
                          </div>
                        </div>''',
                        unsafe_allow_html=True,
                    )

        # Editar e remover eventos
        if eventos:
            st.markdown('<div style="margin-top:20px"></div>', unsafe_allow_html=True)

            tab_ed_ev, tab_rm_ev = st.tabs(['✏️ Editar', '🗑️ Remover'])

            with tab_ed_ev:
                id_edit_ev = st.selectbox(
                    'Evento para editar',
                    options=[e['id'] for e in eventos],
                    format_func=lambda x: next(
                        (e['nome'] for e in eventos if e['id'] == x), str(x)
                    ),
                    key='sel_editar_evento',
                    label_visibility='collapsed',
                )
                ev_sel = next((e for e in eventos if e['id'] == id_edit_ev), None)
                if ev_sel:
                    with st.form('form_editar_evento'):
                        from datetime import datetime as _dte
                        c1e, c2e, c3e = st.columns([3, 1, 2])
                        with c1e:
                            novo_nome_ev = st.text_input('Nome', value=ev_sel.get('nome', ''))
                        with c2e:
                            try:
                                data_ev_atual = _dte.fromisoformat(
                                    str(ev_sel.get('data_evento', ev_sel.get('data', '')))[:10]
                                ).date()
                            except Exception:
                                data_ev_atual = _dte.today().date()
                            nova_data_ev = st.date_input('Data', value=data_ev_atual)
                        with c3e:
                            tipos_ev = ['📝 Prova', '📌 Entrega', '🎯 Apresentação', '📖 Outro']
                            tipo_atual = ev_sel.get('tipo', '📖 Outro')
                            idx_tipo   = tipos_ev.index(tipo_atual) if tipo_atual in tipos_ev else 0
                            novo_tipo_ev = st.selectbox('Tipo', tipos_ev, index=idx_tipo)
                        if st.form_submit_button('💾 Salvar', type='primary'):
                            res = api_patch('eventos', id_edit_ev, {
                                'nome':        novo_nome_ev,
                                'data_evento': nova_data_ev.isoformat(),
                                'tipo':        novo_tipo_ev,
                            })
                            if res and res.status_code in (200, 201, 204):
                                st.success('Evento atualizado!')
                                _invalidar_cache()
                                st.rerun()
                            else:
                                st.error('Erro ao atualizar evento.')

            with tab_rm_ev:
                id_del = st.selectbox(
                    'Evento para remover',
                    options=[e['id'] for e in eventos],
                    format_func=lambda x: next(
                        (e['nome'] for e in eventos if e['id'] == x), str(x)
                    ),
                    key='sel_remover_evento',
                    label_visibility='collapsed',
                )
                if st.button('🗑️ Remover', type='primary'):
                    try:
                        res = requests.delete(
                            f'{BASE_URL}/eventos/{id_del}',
                            headers=get_headers(),
                            timeout=10,
                        )
                        if res and res.status_code in (200, 204):
                            st.success('Evento removido.')
                            _invalidar_cache()
                            st.rerun()
                        else:
                            st.error('Erro ao remover evento.')
                    except Exception as e:
                        st.error(f'Erro: {e}')


# ------------------------------------------------
# 10. MÓDULO RELATÓRIO (separado do dashboard)
# ------------------------------------------------

def modulo_relatorio():
    st.header('📄 Relatório')
    st.caption('Baixe um PDF completo com suas notas e métricas')

    tarefas = api_get('tarefas', st.session_state.get('auth_token', ''))
    discs   = api_get('disciplinas', st.session_state.get('auth_token', ''))

    if not tarefas or not discs:
        st.info('Cadastre atividades para gerar o relatório.')
        return

    df_t = pd.DataFrame(tarefas)
    df_d = pd.DataFrame(discs)

    if 'concluida' not in df_t.columns:
        df_t['concluida'] = False
    df_t['concluida'] = df_t['concluida'].fillna(False).astype(bool)

    nomes_disc = {d['id']: d['nome'] for d in discs}
    df_t['disciplina'] = df_t['disc_id'].map(nomes_disc).fillna('—')

    total_disc    = len(df_d)
    total_tarefas = len(df_t)
    concluidas    = int(df_t['concluida'].sum())
    pendentes     = total_tarefas - concluidas
    media_geral   = float(df_t['nota'].mean()) if 'nota' in df_t.columns else 0.0

    # Preview das métricas antes de baixar
    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric('Disciplinas',  total_disc)
    col2.metric('Atividades',   total_tarefas)
    col3.metric('Concluídas',   concluidas)
    col4.metric('Pendentes',    pendentes)
    col5.metric('Média Geral',  f'{media_geral:.1f}')

    st.markdown('---')
    st.markdown('O relatório inclui: resumo de métricas, tabela de notas com cor por faixa e data de geração.')

    metricas = {
        'total_disciplinas': total_disc,
        'total_tarefas':     total_tarefas,
        'concluidas':        concluidas,
        'pendentes':         pendentes,
        'media_geral':       media_geral,
        'nome_usuario':      st.session_state.get('user_name', 'Estudante'),
    }

    if st.button('📄 Gerar Relatório PDF', type='primary'):
        with st.spinner('Gerando PDF...'):
            eventos_pdf   = api_get('eventos', st.session_state.get('auth_token', ''))   or []
            metas_pdf     = api_get('metas', st.session_state.get('auth_token', ''))      or []
            anotacoes_pdf = api_get('anotacoes', st.session_state.get('auth_token', ''))  or []
            pdf_bytes     = gerar_pdf(df_t, metricas,
                                      eventos=eventos_pdf,
                                      metas=metas_pdf,
                                      anotacoes=anotacoes_pdf)
        st.download_button(
            label='⬇️ Baixar PDF',
            data=pdf_bytes,
            file_name=f"edutrack_relatorio_{date.today().isoformat()}.pdf",
            mime='application/pdf',
        )


# ================================================
# 9. ESTRUTURA PRINCIPAL DE NAVEGAÇÃO
# ================================================
# ATENÇÃO: st.set_page_config() DEVE ser a primeira
# chamada Streamlit do arquivo, antes de qualquer
# outro st.* — por isso fica aqui no final do arquivo
# mas é executada primeiro pelo Python.
# ================================================


# ================================================
# FUNCIONALIDADES MÉDIAS
# ================================================

# ------------------------------------------------
# FEATURE: Alertas de vencimento no dashboard
# Exibidos no topo do dashboard automaticamente
# ------------------------------------------------

def alertas_vencimento(tarefas: list, discs: list):
    '''
    Verifica todas as tarefas com prazo definido e exibe alertas
    no topo do dashboard para as que vencem em até 3 dias.
    Filtra tarefas já concluídas (não faz sentido alertar).
    Cores: vermelho = hoje/atrasado | amarelo = próximos 3 dias.
    '''
    from datetime import datetime as _dt
    if not tarefas:
        return
    nomes_disc = {d['id']: d['nome'] for d in discs}
    alertas = []
    for t in tarefas:
        prazo = t.get('prazo')
        if not prazo or str(prazo) in ('None', 'nan', ''):
            continue
        if t.get('concluida'):
            continue
        try:
            prazo_dt = _dt.fromisoformat(str(prazo)[:10])
            dias = (prazo_dt.date() - _dt.today().date()).days
            if dias <= 3:
                disc_nome = nomes_disc.get(t.get('disc_id'), '—')
                alertas.append((dias, t.get('nome', ''), disc_nome))
        except Exception:
            pass

    if not alertas:
        return

    alertas.sort(key=lambda x: x[0])
    st.markdown('### ⚠️ Atenção — prazos próximos')
    for dias, nome, disc in alertas:
        if dias < 0:
            msg = f'**{nome}** ({disc}) — atrasado {abs(dias)} dia(s)!'
            st.error(msg)
        elif dias == 0:
            st.error(f'**{nome}** ({disc}) — vence **hoje**!')
        elif dias == 1:
            st.warning(f'**{nome}** ({disc}) — vence **amanhã**')
        else:
            st.warning(f'**{nome}** ({disc}) — vence em **{dias} dias**')
    st.markdown('---')


# ------------------------------------------------
# FEATURE: Evolução temporal de notas (gráfico de linha)
# ------------------------------------------------

def modulo_evolucao(df_t: pd.DataFrame, df_d: pd.DataFrame,
                    cor_texto: str, cor_titulo: str, cor_grid: str,
                    template: str = 'plotly_dark', fonte: str = 'Inter, sans-serif',
                    cor_hover_bg: str = '#1e2130', cor_hover_tx: str = '#ffffff'):
    '''Gráfico de linha mostrando evolução das notas por ordem de cadastro.'''
    import plotly.graph_objects as go
    if df_t.empty or 'nota' not in df_t.columns or len(df_t) < 2:
        return

    CORES  = ['#f5c842', '#f97316', '#06b6d4', '#8b5cf6', '#ec4899', '#10b981']
    FONTE  = fonte
    FUNDO  = 'rgba(0,0,0,0)'
    GRID   = cor_grid

    nomes_disc = {d['id']: d['nome'] for d in df_d.to_dict('records')}
    df_ev = df_t.copy()
    df_ev['disciplina'] = df_ev['disc_id'].map(nomes_disc).fillna('—')
    df_ev = df_ev.reset_index(drop=True)
    df_ev['ordem'] = df_ev.index + 1

    fig_linha = go.Figure()

    for i, disc in enumerate(df_ev['disciplina'].unique()):
        df_disc = df_ev[df_ev['disciplina'] == disc].sort_values('ordem')
        cor     = CORES[i % len(CORES)]
        fig_linha.add_trace(go.Scatter(
            x=df_disc['ordem'],
            y=df_disc['nota'],
            name=disc,
            mode='lines+markers',
            line=dict(color=cor, width=2.5, shape='spline', smoothing=0.8),
            marker=dict(
                color='#ffffff',
                size=9,
                line=dict(color=cor, width=2.5),
            ),
            hovertemplate=(
                f'<b>{disc}</b><br>'
                'Atividade nº %{x}<br>'
                'Nota: <b>%{y:.1f}</b><extra></extra>'
            ),
        ))

    # Linha de referência da média mínima (ex: 7.0)
    fig_linha.add_hline(
        y=7.0,
        line_dash='dot',
        line_color='#9ca3af',
        line_width=1,
        annotation_text='  Mínimo aprovação (7.0)',
        annotation_font=dict(size=10, color='#9ca3af'),
        annotation_position='right',
    )

    # Linha de aprovação (7.0)
    fig_linha.add_hrect(
        y0=7.0, y1=10.5,
        fillcolor='rgba(52,211,153,0.04)',
        line_width=0,
    )

    fig_linha.update_layout(
        template=template,
        title=dict(
            text='<b>Evolução das notas ao longo do tempo</b>',
            font=dict(family=FONTE, size=14, color=cor_titulo),
            x=0, xanchor='left', pad=dict(l=4, b=8),
        ),
        plot_bgcolor=FUNDO,
        paper_bgcolor=FUNDO,
        font=dict(family=FONTE, size=11, color=cor_texto),
        height=320,  # mais alto pois ocupa largura total
        margin=dict(l=4, r=4, t=48, b=4),
        xaxis=dict(
            title=dict(text='', font=dict(size=11, color=cor_texto)),
            tickfont=dict(size=11, color=cor_texto, family=FONTE),
            gridcolor=GRID, gridwidth=1,
            linecolor='rgba(255,255,255,0.08)',
            zeroline=False, dtick=1, showgrid=False,
        ),
        yaxis=dict(
            title=dict(text='Nota', font=dict(size=11, color=cor_texto)),
            tickfont=dict(size=11, color=cor_texto, family=FONTE),
            gridcolor=GRID, gridwidth=1,
            linecolor='rgba(255,255,255,0.08)',
            zeroline=False, range=[0, 11], showgrid=True,
            tickvals=[0, 2, 4, 5, 6, 7, 8, 9, 10],
        ),
        legend=dict(
            bgcolor=FUNDO,
            font=dict(family=FONTE, size=11, color=cor_texto),
            orientation='h',
            yanchor='bottom', y=1.01,
            xanchor='right',  x=1,
            title_text='',
        ),
        hoverlabel=dict(
            bgcolor=cor_hover_bg,
            bordercolor='rgba(255,255,255,0.1)',
            font=dict(family=FONTE, size=12, color=cor_hover_tx),
        ),
        hovermode='x unified',
    )
    st.plotly_chart(fig_linha, use_container_width=True)


# ------------------------------------------------
# FEATURE: Metas de desempenho por disciplina
# ------------------------------------------------

def modulo_metas():
    '''
    Metas de desempenho por disciplina.
    O aluno define uma meta de média (ex: 7.0) para cada disciplina.
    O sistema calcula a média atual e mostra uma barra de progresso.
    As metas são salvas no Xano (tabela metas) via POST/PATCH,
    persistindo entre sessões — diferente do session_state que some ao fechar.
    '''
    st.header('🎯 Metas de Desempenho')
    st.caption('Defina uma meta de média para cada disciplina e acompanhe o progresso')

    discs   = api_get('disciplinas', st.session_state.get('auth_token', ''))
    tarefas = api_get('tarefas', st.session_state.get('auth_token', ''))
    metas   = api_get('metas', st.session_state.get('auth_token', '')) or []

    if not discs:
        st.info('Cadastre disciplinas para definir metas.')
        return

    df_t = pd.DataFrame(tarefas) if tarefas else pd.DataFrame()

    # Monta dicionário disc_id → meta do Xano
    metas_xano   = {m['disc_id']: m for m in metas}
    discs_ids    = {d['id'] for d in discs}

    # Remove metas órfãs (disciplina foi deletada)
    for m in metas:
        if m['disc_id'] not in discs_ids:
            api_patch('metas', m['id'], {'disc_id': 0})  # marca como órfã

    for disc in discs:
        disc_id   = disc['id']
        disc_nome = disc['nome']

        # Calcula média atual
        if not df_t.empty and 'disc_id' in df_t.columns:
            notas = df_t[df_t['disc_id'] == disc_id]['nota'].dropna()
            media_atual = float(notas.mean()) if len(notas) > 0 else 0.0
            qtd = len(notas)
        else:
            media_atual = 0.0
            qtd = 0

        # Meta do Xano ou padrão 7.0
        meta_registro = metas_xano.get(disc_id)
        meta_valor    = float(meta_registro['valor_meta']) if meta_registro else 7.0

        with st.container():
            col1, col2, col3 = st.columns([3, 1, 1])
            with col1:
                st.markdown(f'**{disc_nome}**')
                progresso = min(media_atual / meta_valor, 1.0) if meta_valor > 0 else 0
                pct = int(progresso * 100)
                st.progress(progresso)
                st.caption(f'{pct}% da meta ({qtd} atividade(s))')
            with col2:
                st.metric('Média atual', f'{media_atual:.1f}')
            with col3:
                nova_meta = st.number_input(
                    'Meta',
                    min_value=0.0, max_value=10.0,
                    value=meta_valor, step=0.5,
                    key=f'meta_{disc_id}',
                )

            col_s1, col_s2 = st.columns([1, 4])
            with col_s1:
                if st.button('💾 Salvar meta', key=f'btn_meta_{disc_id}', type='primary'):
                    if meta_registro:
                        # Atualiza meta existente
                        res = api_patch('metas', meta_registro['id'], {'valor_meta': nova_meta})
                    else:
                        # Cria nova meta
                        res = api_post('metas', {'disc_id': disc_id, 'valor_meta': nova_meta})
                    if res and res.status_code in (200, 201, 204):
                        st.success('Meta salva!')
                        _invalidar_cache()
                        st.rerun()
                    else:
                        st.error('Erro ao salvar meta.')

            # Status visual
            if qtd == 0:
                st.caption('Sem atividades ainda')
            elif media_atual >= nova_meta:
                st.success(f'✅ Meta atingida! ({media_atual:.1f} ≥ {nova_meta:.1f})')
            else:
                falta = nova_meta - media_atual
                st.warning(f'Faltam {falta:.1f} pontos para atingir a meta')

            st.markdown('---')


# ------------------------------------------------
# FEATURE: Perfil do aluno
# ------------------------------------------------

def modulo_perfil():
    '''
    Perfil do aluno com dados persistentes no Xano.
    Seções:
      1. Identidade — avatar com iniciais, upload de foto,
         campos editáveis (nome, curso, instituição, semestre)
         salvos via PATCH /auth/me
      2. Estatísticas — métricas calculadas das tarefas
      3. Progresso por disciplina — barras coloridas por média
      4. Conta — botão de logout
    '''
    st.header('👤 Meu Perfil')

    # ── Dados — busca perfil atualizado do Xano ──────────
    me_data = None
    try:
        me_res = requests.get(
            f'{BASE_URL}/auth/me',
            headers=get_headers(), timeout=10,
        )
        if me_res.status_code == 200:
            me_data = me_res.json()
            st.session_state.user_name   = me_data.get('name',        st.session_state.get('user_name', 'Estudante'))
            st.session_state.curso       = me_data.get('curso',       st.session_state.get('curso', ''))
            st.session_state.instituicao = me_data.get('instituicao', st.session_state.get('instituicao', ''))
            st.session_state.semestre    = me_data.get('semestre',    st.session_state.get('semestre', ''))
    except Exception:
        pass

    nome    = st.session_state.get('user_name', 'Estudante')
    tarefas = api_get('tarefas', st.session_state.get('auth_token', ''))
    discs   = api_get('disciplinas', st.session_state.get('auth_token', ''))
    profs   = api_get('professores', st.session_state.get('auth_token', ''))

    df_t = pd.DataFrame(tarefas) if tarefas else pd.DataFrame()
    if not df_t.empty and 'concluida' in df_t.columns:
        df_t['concluida'] = df_t['concluida'].fillna(False).astype(bool)
        concluidas  = int(df_t['concluida'].sum())
        media_geral = float(df_t['nota'].mean()) if 'nota' in df_t.columns else 0.0
        melhor_nota = float(df_t['nota'].max())  if 'nota' in df_t.columns else 0.0
    else:
        concluidas = media_geral = melhor_nota = 0.0

    total_disc = len(discs) if discs else 0
    total_ativ = len(df_t)
    iniciais   = ''.join([p[0].upper() for p in nome.split()[:2]]) if nome else 'EU'

    # ── SEÇÃO 1: Identidade ────────────────────────────
    st.subheader('Identidade')

    col_av, col_info = st.columns([1, 3])

    with col_av:
        # Avatar com iniciais e opção de foto
        foto_atual = st.session_state.get('foto_perfil', None)
        if foto_atual:
            st.image(foto_atual, width=100, caption='Foto de perfil')
        else:
            st.markdown(
                f'<div style="width:90px;height:90px;border-radius:50%;'
                f'background:#f5c842;display:flex;align-items:center;'
                f'justify-content:center;font-size:28px;font-weight:500;'
                f'color:#6b4f00;margin-bottom:8px">{iniciais}</div>',
                unsafe_allow_html=True,
            )


    with col_info:
        with st.form('form_perfil'):
            c1, c2 = st.columns(2)
            with c1:
                novo_nome = st.text_input('Nome completo', value=nome)
                curso     = st.text_input('Curso',
                                          value=st.session_state.get('curso', ''),
                                          placeholder='Ex: Engenharia de Software')
            with c2:
                instituicao = st.text_input('Instituição',
                                            value=st.session_state.get('instituicao', ''),
                                            placeholder='Ex: FIAP')
                semestre    = st.text_input('Semestre',
                                            value=st.session_state.get('semestre', ''),
                                            placeholder='Ex: 3º semestre')
            if st.form_submit_button('💾 Salvar alterações', type='primary'):
                # Salva no Xano via PATCH /auth/me
                payload = {
                    'name':        novo_nome,
                    'curso':       curso,
                    'instituicao': instituicao,
                    'semestre':    semestre,
                }
                try:
                    res = requests.patch(
                        f'{BASE_URL}/auth/me',
                        json=payload,
                        headers=get_headers(),
                        timeout=10,
                    )
                    if res.status_code in (200, 201, 204):
                        st.session_state.user_name   = novo_nome
                        st.session_state.curso       = curso
                        st.session_state.instituicao = instituicao
                        st.session_state.semestre    = semestre
                        st.success('✅ Perfil salvo com sucesso!')
                    else:
                        st.error(f'Erro ao salvar: {res.status_code} — {res.text[:100]}')
                except Exception as e:
                    st.error(f'Erro de conexão: {e}')
                st.rerun()

    st.markdown('---')

    # ── SEÇÃO 2: Estatísticas gerais ──────────────────
    st.subheader('Estatísticas gerais')

    col1, col2, col3, col4 = st.columns(4)
    col1.metric('Disciplinas',  total_disc,       help='Total de disciplinas cadastradas')
    col2.metric('Atividades',   total_ativ,        help='Total de atividades lançadas')
    col3.metric('Concluídas',   int(concluidas),   help='Atividades marcadas como concluídas')
    col4.metric('Média geral',  f'{media_geral:.1f}', help='Média de todas as notas')

    st.markdown('---')

    # ── SEÇÃO 3: Progresso por disciplina ─────────────
    st.subheader('Progresso por disciplina')

    if not discs:
        st.info('Nenhuma disciplina cadastrada ainda.')
    else:
        nomes_disc = {d['id']: d['nome'] for d in discs}
        for disc in discs:
            disc_id   = disc['id']
            disc_nome = disc['nome']
            if not df_t.empty and 'disc_id' in df_t.columns:
                notas = df_t[df_t['disc_id'] == disc_id]['nota'].dropna()
                media = float(notas.mean()) if len(notas) > 0 else 0.0
                qtd   = len(notas)
            else:
                media = 0.0
                qtd   = 0

            # Cor da barra por faixa de nota
            cor = '#34d399' if media >= 7 else '#f5c842' if media >= 5 else '#f87171'
            pct = int((media / 10) * 100)

            col_nome, col_barra, col_nota = st.columns([2, 4, 1])
            with col_nome:
                st.write(f'**{disc_nome}**')
                st.caption(f'{qtd} atividade(s)')
            with col_barra:
                st.markdown(
                    f'<div style="margin-top:10px;height:8px;border-radius:4px;'
                    f'background:#e5e7eb;overflow:hidden">'
                    f'<div style="width:{pct}%;height:100%;background:{cor};'
                    f'border-radius:4px"></div></div>',
                    unsafe_allow_html=True,
                )
            with col_nota:
                st.metric('', f'{media:.1f}')

    st.markdown('---')

    # ── SEÇÃO 4: Conta ────────────────────────────────
    st.subheader('Conta')
    st.caption('Encerre sua sessão atual')
    if st.button('Sair da conta', type='primary'):
        st.session_state.clear()
        _invalidar_cache()
        st.rerun()


# ------------------------------------------------
# FEATURE: Anotações por disciplina
# ------------------------------------------------

def modulo_anotacoes():
    '''
    Anotações livres por disciplina.
    O aluno pode salvar resumos, links e observações sobre cada matéria.
    As anotações são salvas no Xano (tabela anotacoes) — uma por disciplina.
    Se já existe anotação para a disciplina, faz PATCH (atualiza).
    Se não existe, faz POST (cria nova).
    '''
    st.header('📝 Anotações')
    st.caption('Salve resumos, links e anotações por disciplina')

    discs     = api_get('disciplinas', st.session_state.get('auth_token', ''))
    anotacoes = api_get('anotacoes', st.session_state.get('auth_token', '')) or []

    if not discs:
        st.info('Cadastre disciplinas para criar anotações.')
        return

    disc_nomes = {d['id']: d['nome'] for d in discs}

    # Monta dicionário disc_id → registro do Xano
    discs_ids_an = {d['id'] for d in discs}
    anot_xano    = {
        a['disc_id']: a for a in anotacoes
        if a['disc_id'] in discs_ids_an  # ignora anotações órfãs
    }

    disc_sel = st.selectbox(
        'Selecione a disciplina',
        options=[d['id'] for d in discs],
        format_func=lambda x: disc_nomes.get(x, str(x)),
    )

    anot_registro = anot_xano.get(disc_sel)
    texto_atual   = anot_registro['texto'] if anot_registro else ''

    st.markdown(f'**{disc_nomes.get(disc_sel, "")}** — anotações livres')
    novo_texto = st.text_area(
        'Escreva suas anotações, resumos ou links aqui...',
        value=texto_atual,
        height=300,
        label_visibility='collapsed',
    )

    col1, col2 = st.columns([1, 4])
    with col1:
        if st.button('💾 Salvar', type='primary'):
            if anot_registro:
                res = api_patch('anotacoes', anot_registro['id'], {'texto': novo_texto})
            else:
                res = api_post('anotacoes', {'disc_id': disc_sel, 'texto': novo_texto})
            if res and res.status_code in (200, 201, 204):
                st.success('Anotação salva!')
                _invalidar_cache()
                st.rerun()
            else:
                st.error('Erro ao salvar anotação.')
    with col2:
        if anot_registro and st.button('🗑️ Limpar'):
            res = api_patch('anotacoes', anot_registro['id'], {'texto': ''})
            if res and res.status_code in (200, 201, 204):
                _invalidar_cache()
                st.rerun()

    # Lista todas as anotações com conteúdo
    anot_com_texto = [a for a in anotacoes if a.get('texto', '').strip()]
    if anot_com_texto:
        st.markdown('---')
        st.subheader('Todas as anotações')
        for anot in anot_com_texto:
            nome_disc = disc_nomes.get(anot['disc_id'], str(anot['disc_id']))
            with st.expander(f'📖 {nome_disc}'):
                st.markdown(anot['texto'])


# ------------------------------------------------
# FEATURE: Histórico de conquistas (timeline)
# ------------------------------------------------

def modulo_historico():
    '''
    Histórico de conquistas — linha do tempo das atividades concluídas.
    Filtra as tarefas com concluida=True e as exibe em ordem,
    com badge colorido por faixa de nota:
      🥇 Ouro (≥9.0) | ✅ Verde (≥7.0) | 📌 Amarelo (≥5.0) | ⚠️ Vermelho (<5.0)
    Dados vêm do Xano em tempo real — não requer tabela extra.
    '''
    st.header('🏆 Histórico de Conquistas')
    st.caption('Linha do tempo das atividades concluídas')

    tarefas = api_get('tarefas', st.session_state.get('auth_token', ''))
    discs   = api_get('disciplinas', st.session_state.get('auth_token', ''))

    if not tarefas or not discs:
        st.info('Conclua atividades para ver seu histórico.')
        return

    df_t = pd.DataFrame(tarefas)
    if 'concluida' not in df_t.columns:
        df_t['concluida'] = False
    df_t['concluida'] = df_t['concluida'].fillna(False).astype(bool)

    df_concluidas = df_t[df_t['concluida'] == True].copy()
    if df_concluidas.empty:
        st.info('Nenhuma atividade concluída ainda. Marque atividades como concluídas na aba Tarefas/Notas.')
        return

    nomes_disc = {d['id']: d['nome'] for d in discs}
    df_concluidas['disciplina'] = df_concluidas['disc_id'].map(nomes_disc).fillna('—')

    # Estatísticas de conquistas
    total_conc  = len(df_concluidas)
    media_conc  = float(df_concluidas['nota'].mean()) if 'nota' in df_concluidas.columns else 0.0
    notas_10    = len(df_concluidas[df_concluidas['nota'] == 10.0]) if 'nota' in df_concluidas.columns else 0

    col1, col2, col3 = st.columns(3)
    col1.metric('Atividades concluídas', total_conc)
    col2.metric('Média das concluídas',  f'{media_conc:.1f}')
    col3.metric('Notas 10.0',            notas_10)

    st.markdown('---')

    # Timeline visual
    for _, row in df_concluidas.iterrows():
        nota_val = row.get('nota', 0)
        disc     = row.get('disciplina', '—')
        nome     = row.get('nome', '')

        if nota_val >= 9:
            cor_badge = ('#fef9e7', '#92400e', '🥇')
        elif nota_val >= 7:
            cor_badge = ('#dcfce7', '#166534', '✅')
        elif nota_val >= 5:
            cor_badge = ('#fef3c7', '#92400e', '📌')
        else:
            cor_badge = ('#fee2e2', '#991b1b', '⚠️')

        bg, txt, icone = cor_badge
        col_a, col_b, col_c = st.columns([3, 2, 1])
        with col_a:
            st.write(f'{icone} **{nome}**')
        with col_b:
            st.caption(disc)
        with col_c:
            st.metric('', f'{nota_val:.1f}')



# ------------------------------------------------
# MÓDULO: Meu Espaço (Metas + Anotações + Histórico)
# ------------------------------------------------

def modulo_meu_espaco():
    st.header('🗂️ Meu Espaço')
    st.caption('Metas de desempenho, anotações por disciplina e histórico de conquistas')

    aba_metas, aba_anotacoes, aba_historico = st.tabs([
        '🎯  Metas',
        '📝  Anotações',
        '🏆  Histórico',
    ])

    with aba_metas:
        modulo_metas()

    with aba_anotacoes:
        modulo_anotacoes()

    with aba_historico:
        modulo_historico()


st.set_page_config(
    page_title='EduTrack AI',
    layout='wide',
    page_icon='📚',
    initial_sidebar_state='expanded',  # sidebar sempre aberta
)

# ================================================
# AQUI É ONDE O CSS É APLICADO
# ================================================
# Chamamos aplicar_css() logo após set_page_config().
# Isso garante que o CSS seja injetado ANTES de
# qualquer componente visual ser renderizado.
# Se você mover essa linha para depois de st.title()
# ou st.header(), o CSS ainda funciona, mas você
# pode ver um "flash" do estilo padrão antes do custom.
# ================================================
aplicar_css()  # <--- INJEÇÃO DO CSS ACONTECE AQUI

# ================================================
# 9. NAVEGAÇÃO PRINCIPAL
# ================================================
# Controla qual tela exibir baseado no estado de autenticação.
# Fluxo:
#   1. Verifica se há token na URL (?token=xxx)
#   2. Se sim, valida o token com /auth/me e restaura a sessão
#   3. Se não estiver logado, exibe a tela de login
#   4. Se estiver logado, exibe a sidebar e o módulo selecionado
# ================================================

if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False

# Restaura sessão via token salvo na URL (sobrevive ao refresh)
if not st.session_state.logged_in:
    token_url = st.query_params.get('token', '')
    if token_url:
        try:
            me_v = requests.get(
                f'{BASE_URL}/auth/me',
                headers={'Authorization': f'Bearer {token_url}'},
                timeout=10,
            )
            if me_v.status_code == 200:
                me_vd = me_v.json()
                st.session_state.auth_token  = token_url
                st.session_state.logged_in   = True
                st.session_state.user_name   = me_vd.get('name', me_vd.get('email', ''))
                st.session_state.curso       = me_vd.get('curso', '')
                st.session_state.instituicao = me_vd.get('instituicao', '')
                st.session_state.semestre    = me_vd.get('semestre', '')
                st.rerun()
        except Exception:
            pass

if not st.session_state.logged_in:
    st.query_params.clear()
    tela_acesso()
else:
    with st.sidebar:
        st.title('📚 EduTrack AI')
        _opcoes_menu = ['Dashboard', 'Professores', 'Disciplinas', 'Tarefas/Notas',
                       'Calendário', 'Meu Espaço', 'Perfil']

        # Navegação programática via _nav_target
        # Guarda em pagina_atual para persistir entre reruns
        if '_nav_target' in st.session_state:
            st.session_state.pagina_atual = st.session_state.pop('_nav_target')

        if 'pagina_atual' not in st.session_state:
            st.session_state.pagina_atual = 'Dashboard'

        # Radio sem key — não interfere com o estado da sidebar
        _idx = _opcoes_menu.index(st.session_state.pagina_atual)                if st.session_state.pagina_atual in _opcoes_menu else 0
        menu = st.radio('Menu', _opcoes_menu, index=_idx)

        # Sincroniza seleção manual do usuário
        st.session_state.pagina_atual = menu
        st.markdown('---')
        if st.button('Sair', use_container_width=True):
            st.session_state.clear()
            _invalidar_cache()
            st.query_params.clear()
            st.rerun()

    # Roteamento: cada opção do menu chama o módulo correspondente
    # O 'match' é equivalente a um switch/case — mais eficiente que múltiplos if/elif



    match menu:
        case 'Dashboard':       modulo_dashboard()
        case 'Professores':     modulo_professores()
        case 'Disciplinas':     modulo_disciplinas()
        case 'Tarefas/Notas':   modulo_tarefas()
        case 'Calendário':      modulo_calendario()
        case 'Meu Espaço':      modulo_meu_espaco()
        case 'Perfil':          modulo_perfil()